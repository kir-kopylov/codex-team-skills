#!/usr/bin/env python3
"""Нормализует CSV/XLSX-выгрузку 1С для отчета по марже квартир."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

OUTPUT_FIELDS = [
    "source_sheet",
    "source_row",
    "operation_date",
    "apartment",
    "expense_category",
    "amount",
    "attribution_type",
    "allocation_base",
    "account_debit",
    "account_credit",
    "counterparty",
    "contract",
    "document",
    "description",
    "needs_review",
    "comment",
]

HEADER_ALIASES = {
    "operation_date": ["дата", "период"],
    "account_debit": ["счет дт", "дебет", "дт", "счет дебета"],
    "account_credit": ["счет кт", "кредит", "кт", "счет кредита"],
    "counterparty": ["контрагент", "поставщик", "арендодатель"],
    "contract": ["договор"],
    "document": ["документ", "регистратор", "документ-регистратор"],
    "apartment": ["адрес", "квартира", "объект", "помещение", "подразделение", "номенклатурная группа"],
    "description": ["содержание", "операция", "назначение", "комментарий", "статья затрат", "статья расходов"],
    "amount": ["сумма", "оборот", "расход", "затраты", "стоимость"],
}

CLASSIFIERS = [
    ("Аренда", ["аренд", "собственник"], "direct", ""),
    ("Коммунальные услуги", ["коммун", "жкх", "электро", "водоснаб", "отоплен"], "direct", ""),
    ("Интернет и связь", ["интернет", "связь", "домофон", "телевид"], "direct", ""),
    ("Ремонт и улучшения", ["ремонт", "улучш", "мебел", "техник", "замок", "мастер", "матрас"], "direct", ""),
    ("Уборка", ["уборк", "клининг", "горнич"], "direct", ""),
    ("Стирка и прачечная", ["стирк", "прач", "химчист"], "allocated", "уборки или бронирования"),
    ("Расходники", ["бумаг", "сахар", "чай", "кофе", "мыло", "химия", "расходник"], "allocated", "бронирования или ночи"),
    ("Белье и текстиль", ["бель", "полотен", "простын", "текстил"], "allocated", "уборки или спальные места"),
    ("Зарплата персонала", ["зарплат", "зп", "оклад", "сотрудник", "персонал", "администратор"], "allocated", "выручка или бронирования"),
    ("Комиссии", ["комисс", "эквайр", "банк", "авито", "суточно", "островок"], "direct", ""),
]

TOTAL_MARKERS = ("итого", "общий итог", "начальное сальдо", "конечное сальдо", "оборот за")


@dataclass
class SheetRows:
    name: str
    rows: list[list[Any]]


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u00a0", " ").replace("ё", "е").lower()
    return re.sub(r"\s+", " ", text).strip()


def compact(value: Any) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", norm(value))


def parse_amount(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, int | float | Decimal):
        return Decimal(str(value))
    text = str(value).replace("\u00a0", " ").strip()
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")") or text.endswith("-")
    text = text.strip("()")
    if text.endswith("-"):
        text = text[:-1]
    text = re.sub(r"(?i)(руб\.?|₽)", "", text)
    text = re.sub(r"[^0-9,.-]", "", text.replace(" ", ""))
    if not text or text in {"-", ",", "."}:
        return None
    if "," in text and "." in text:
        decimal_sep = "," if text.rfind(",") > text.rfind(".") else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "").replace(decimal_sep, ".")
    else:
        text = text.replace(",", ".")
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def read_csv(path: Path) -> list[SheetRows]:
    for encoding in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            raw = path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        try:
            dialect = csv.Sniffer().sniff(raw[:4096], delimiters=";\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return [SheetRows(path.stem, list(csv.reader(raw.splitlines(), dialect)))]
    raise RuntimeError("Не удалось прочитать CSV: неизвестная кодировка.")


def read_xlsx(path: Path) -> list[SheetRows]:
    if openpyxl is None:
        raise RuntimeError("Для XLSX нужен пакет openpyxl. Сохраните файл как CSV или установите openpyxl.")
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return [SheetRows(sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]) for sheet in workbook.worksheets]


def load_input(path: Path) -> list[SheetRows]:
    if path.suffix.lower() in {".csv", ".txt"}:
        return read_csv(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise RuntimeError("Поддерживаются только CSV, TXT, XLSX и XLSM.")


def detect_header(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        joined = " | ".join(norm(cell) for cell in row)
        score = sum(1 for aliases in HEADER_ALIASES.values() if any(alias in joined for alias in aliases))
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score >= 2 else 0


def map_columns(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [compact(header) for header in headers]
    for field, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if any(compact(alias) in header for alias in aliases):
                mapping[field] = index
                break
    return mapping


def cell(row: list[Any], mapping: dict[str, int], field: str) -> str:
    index = mapping.get(field)
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def row_text(row: list[Any]) -> str:
    return " | ".join(str(value) for value in row if value is not None and str(value).strip())


def is_total_or_empty(row: list[Any]) -> bool:
    values = [norm(value) for value in row if value is not None and str(value).strip()]
    if not values:
        return True
    joined = " ".join(values[:4])
    return any(joined.startswith(marker) for marker in TOTAL_MARKERS)


def read_addresses(path: Path | None) -> list[str]:
    if path is None:
        return []
    return [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines() if line.strip()]


def detect_apartment(row: list[Any], mapping: dict[str, int], addresses: list[str]) -> tuple[str, str]:
    text = row_text(row)
    normalized = norm(text)
    for address in addresses:
        if norm(address) in normalized:
            return address, "найдено по списку адресов"
    candidate = cell(row, mapping, "apartment")
    if candidate and any(ch.isdigit() for ch in candidate):
        return candidate, "найдено в аналитике"
    contract = cell(row, mapping, "contract")
    if contract and any(ch.isdigit() for ch in contract):
        return contract, "найдено в договоре"
    return "", "квартира не найдена"


def classify(row: list[Any], mapping: dict[str, int], has_apartment: bool) -> tuple[str, str, str, str]:
    text = norm(" | ".join([cell(row, mapping, "description"), cell(row, mapping, "counterparty"), cell(row, mapping, "contract"), row_text(row)]))
    for category, needles, attribution, base in CLASSIFIERS:
        if any(needle in text for needle in needles):
            if attribution == "direct" and not has_apartment:
                return category, "needs_review", base, "обычно прямой расход, но квартира не найдена"
            if attribution == "allocated" and has_apartment:
                return category, attribution, base, "обычно распределяемый расход; проверьте, не указан ли факт по квартире"
            return category, attribution, base, ""
    return "Прочее / требует классификации", "direct" if has_apartment else "needs_review", "", "категория не распознана"


def normalize_rows(sheets: list[SheetRows], addresses: list[str]) -> tuple[list[dict[str, str]], list[str]]:
    result: list[dict[str, str]] = []
    notes: list[str] = []
    skipped_totals = 0
    skipped_without_amount = 0
    for sheet in sheets:
        if not sheet.rows:
            continue
        header_index = detect_header(sheet.rows)
        mapping = map_columns(sheet.rows[header_index])
        for offset, row in enumerate(sheet.rows[header_index + 1 :], start=header_index + 2):
            if is_total_or_empty(row):
                skipped_totals += 1
                continue
            amount = parse_amount(cell(row, mapping, "amount"))
            if amount is None:
                parsed_values = [parse_amount(value) for value in row]
                amount = next((value for value in reversed(parsed_values) if value is not None), None)
            if amount is None:
                skipped_without_amount += 1
                continue
            apartment, apartment_comment = detect_apartment(row, mapping, addresses)
            category, attribution, base, class_comment = classify(row, mapping, bool(apartment))
            needs_review = attribution == "needs_review" or not apartment or "требует" in category
            comments = "; ".join(part for part in [apartment_comment if not apartment else "", class_comment] if part)
            result.append({
                "source_sheet": sheet.name,
                "source_row": str(offset),
                "operation_date": cell(row, mapping, "operation_date"),
                "apartment": apartment,
                "expense_category": category,
                "amount": str(amount.quantize(Decimal("0.01"))),
                "attribution_type": attribution,
                "allocation_base": base,
                "account_debit": cell(row, mapping, "account_debit"),
                "account_credit": cell(row, mapping, "account_credit"),
                "counterparty": cell(row, mapping, "counterparty"),
                "contract": cell(row, mapping, "contract"),
                "document": cell(row, mapping, "document"),
                "description": cell(row, mapping, "description") or row_text(row),
                "needs_review": "yes" if needs_review else "no",
                "comment": comments,
            })
    notes.append(f"Пропущено служебных или пустых строк: {skipped_totals}")
    notes.append(f"Пропущено строк без суммы: {skipped_without_amount}")
    return result, notes


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def build_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    for row in rows:
        amount = parse_amount(row["amount"]) or Decimal("0")
        grouped[(row["apartment"] or "[без квартиры]", row["expense_category"], row["attribution_type"])] += amount
    return [
        {"apartment": apartment, "expense_category": category, "attribution_type": attribution, "amount": str(amount.quantize(Decimal("0.01")))}
        for (apartment, category, attribution), amount in sorted(grouped.items())
    ]


def write_xlsx(path: Path, rows: list[dict[str, str]], summary: list[dict[str, str]], report: list[str]) -> None:
    if openpyxl is None:
        return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "normalized_expenses"
    sheet.append(OUTPUT_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in OUTPUT_FIELDS])
    summary_sheet = workbook.create_sheet("summary_by_apartment")
    summary_fields = ["apartment", "expense_category", "attribution_type", "amount"]
    summary_sheet.append(summary_fields)
    for row in summary:
        summary_sheet.append([row.get(field, "") for field in summary_fields])
    report_sheet = workbook.create_sheet("data_quality")
    for line in report:
        report_sheet.append([line])
    workbook.save(path)


def build_report(rows: list[dict[str, str]], notes: list[str]) -> list[str]:
    total = sum((parse_amount(row["amount"]) or Decimal("0") for row in rows), Decimal("0"))
    without_apartment = sum(1 for row in rows if not row["apartment"])
    review = sum(1 for row in rows if row["needs_review"] == "yes")
    allocated = sum(1 for row in rows if row["attribution_type"] == "allocated")
    return [
        "Отчет о качестве данных",
        f"Нормализовано строк: {len(rows)}",
        f"Общая сумма: {total.quantize(Decimal('0.01'))}",
        f"Строк без квартиры: {without_apartment}",
        f"Строк требуют проверки: {review}",
        f"Строк распределяемых расходов: {allocated}",
        "",
        "Технические заметки:",
        *[f"- {note}" for note in notes],
        "",
        "Напоминание: стирку, расходники, белье и зарплату показывайте как расчетно распределенные, если поквартирный учет не велся.",
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Нормализовать выгрузку 1С для отчета по марже квартир.")
    parser.add_argument("input", type=Path, help="Файл CSV, TXT, XLSX или XLSM из 1С.")
    parser.add_argument("--output-dir", type=Path, default=Path("out"), help="Папка для результата.")
    parser.add_argument("--addresses", type=Path, help="Необязательный UTF-8 файл: одна квартира или адрес на строку.")
    args = parser.parse_args()

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    rows, notes = normalize_rows(load_input(args.input), read_addresses(args.addresses))
    summary = build_summary(rows)
    report = build_report(rows, notes)

    write_csv(output_dir / "normalized_expenses.csv", rows, OUTPUT_FIELDS)
    write_csv(output_dir / "expense_summary_by_apartment.csv", summary, ["apartment", "expense_category", "attribution_type", "amount"])
    (output_dir / "data_quality_report.txt").write_text("\n".join(report), encoding="utf-8")
    write_xlsx(output_dir / "normalized_expenses.xlsx", rows, summary, report)

    print(f"Нормализовано строк: {len(rows)}")
    print(f"Результат: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
