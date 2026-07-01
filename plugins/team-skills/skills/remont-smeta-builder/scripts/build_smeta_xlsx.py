#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Параметризованный сборщик Excel-сметы на ремонт квартиры.

Не хардкодит регион и НЕ заполняет цены: город/валюта/комнаты/проёмы — на вход,
ячейки цен пустые (заполняются живыми офферами под локацию и текущую дату).

Использование:
    python3 build_smeta_xlsx.py --out /путь/смета.xlsx --city "Астана" --currency ₸
    python3 build_smeta_xlsx.py --out out.xlsx --config config.json

config.json (все поля опциональны, есть дефолты):
{
  "city": "Астана",
  "currency": "₸",
  "title": "Объект А, 2-комн.",
  "rooms":    [{"name": "Зал", "floor": 17.3}, {"name": "Кухня", "floor": 5.3}],
  "openings": [{"element": "Дверь — спальня", "room": "Спальня", "qty": 1}],
  "materials":[{"name": "Плитка напольная", "unit": "м²", "waste": 10}]
}

Книга: «Сводка», «Замеры — проёмы», «Замеры — стены и пол», «Тесты замеров»,
«Решения и проверки», «Смета», «Материалы (ведомость)», «Справочники».

Защита от потери данных: если --out уже существует, нужен флаг --force,
иначе сборка прерывается (книга — единственный носитель внесённых цен/замеров).
"""
from __future__ import annotations
import argparse, json, sys, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------- палитра / стили ----------
NAVY="1F3864"; HEAD="2E75B6"; SECT="D6E4F0"; SUB="EAF1FB"; R2FILL="FCE4D6"
INPUT="FFFDF5"; TOTAL="C6E0B4"; GREY="F2F2F2"; WHITE="FFFFFF"; GHOSTC="A6A6A6"; CALC="EDF3FA"
AR="Arial"
_thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=_thin,right=_thin,top=_thin,bottom=_thin)
def F(sz=10,b=False,color="000000",it=False): return Font(name=AR,size=sz,bold=b,color=color,italic=it)
def FILL(c): return PatternFill("solid",fgColor=c)
def AL(h="left",wrap=True,v="center"): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def SC(c,font=None,fillc=None,align=None,bd=True,numfmt=None):
    c.font=font or F()
    if fillc: c.fill=FILL(fillc)
    c.alignment=align or AL()
    if bd: c.border=BORDER
    if numfmt: c.number_format=numfmt

# ---------- дефолты (обезличенный шаблон, без адреса и без цен) ----------
DEFAULT_ROOMS=[{"name":"Зал","floor":None},{"name":"Спальня","floor":None},{"name":"Кухня","floor":None},
               {"name":"Коридор","floor":None},{"name":"Санузел","floor":None},{"name":"Балкон","floor":None}]
DEFAULT_OPENINGS=[{"element":"Дверь — спальня","room":"Спальня","qty":1},
                  {"element":"Дверь — санузел","room":"Санузел","qty":1},
                  {"element":"Дверь входная","room":"Коридор","qty":1},
                  {"element":"Окно — зал","room":"Зал","qty":1},
                  {"element":"Окно — спальня","room":"Спальня","qty":1},
                  {"element":"Окно — кухня","room":"Кухня","qty":1}]
DEFAULT_MATERIALS=[("Плитка напольная","м²",10),("Плитка настенная","м²",15),("Плиточный клей","меш",None),
 ("Затирка","кг",None),("Ламинат","м²",5),("Подложка","м²",5),("Наливной пол/ровнитель","меш",None),
 ("Пескосмесь (стяжка)","меш",None),("Грунтовка","л",None),("Шпаклёвка старт/финиш","меш",None),
 ("Серпянка/сетка","рул",None),("Краска стен","л",None),("Краска потолков","л",None),
 ("Гидроизоляция (смесь)","кг",None),("Галтель потолочная","м.п.",10),("Плинтус напольный","м.п.",10),
 ("Пороги/стыки","шт",None),("Трубы ВС + фитинги","компл",None),("Трубы канализации","компл",None),
 ("Унитаз","шт",None),("Раковина + смеситель","компл",None),("Смеситель кухня","шт",None),
 ("Ванна / душевая","компл",None),("Полотенцесушитель","шт",None),("Счётчики воды","шт",None),
 ("Краны/фильтр/редуктор","компл",None),("Радиаторы","шт",None),("Кабель","м",None),
 ("Щит + автоматы + УЗО","компл",None),("Розетки/выключатели + рамки","шт",None),
 ("Подрозетники/коробки","шт",None),("Двери межкомнатные + фурнитура","шт",None),("Дверь входная","шт",None),
 ("Доборы/наличники","компл",None),("Остекление балкона (отд. подряд)","компл",None),
 ("Расходники (крепёж, скотч, плёнка, валики)","компл",None)]
# Универсальные виды работ (без цен). None во 2-й позиции = подзаголовок раздела.
R1=[("Демонтаж",None,None,None),
 ("Демонтаж старого пола","м²","Замеры: пол",""),("Демонтаж старой плитки","м²","Замеры: пол",""),
 ("Демонтаж старой сантехники","компл","Решения",""),("Демонтаж старой мебели","компл","осмотр",""),
 ("Демонтаж дверей","шт","Замеры: проёмы",""),("Вынос и вывоз мусора","рейс","расчёт по объёму",""),
 ("Полы",None,None,None),
 ("Грунтовка основания","м²","Замеры: пол",""),("Наливной пол / ровнитель","м²","Замеры: пол",""),
 ("Стяжка в санузле + уклон","м²","Замеры: пол",""),("Укладка чистового пола + подложка","м²","Замеры: пол",""),
 ("Плинтус напольный","м.п.","Замеры: периметр",""),("Пороги / стыки","шт","Решения",""),
 ("Стены",None,None,None),
 ("Счистка покрытия + смывка","м²","Замеры: стены нетто",""),("Заделка трещин + серпянка","м.п.","осмотр",""),
 ("Грунтовка стен","м²","Замеры: стены нетто",""),("Шпаклёвка под покраску + шлифовка","м²","Замеры: стены нетто",""),
 ("Откосы окон/дверей","м.п.","Замеры: проёмы",""),("Покраска стен (2 слоя)","м²","Замеры: стены нетто",""),
 ("Потолки",None,None,None),
 ("Смывка побелки","м²","Замеры: пол=потолок",""),("Шпаклёвка потолков + шлифовка","м²","Замеры: пол=потолок",""),
 ("Грунт + покраска потолков (2 слоя)","м²","Замеры: пол=потолок",""),("Монтаж галтели + покраска","м.п.","Замеры: периметр",""),
 ("Плитка и санузел",None,None,None),
 ("Гидроизоляция санузла","м²","Замеры: санузел",""),("Укладка плитки на пол","м²","Замеры: санузел/кухня",""),
 ("Укладка плитки на стены","м²","Замеры: санузел",""),("Фартук на кухне","м²","Замеры: кухня",""),
 ("Затирка, запил, уголки","компл","",""),("Короб на стояк + лючок","компл","Решения",""),
 ("Вытяжной вентилятор","шт","",""),
 ("Сантехника",None,None,None),
 ("Разводка водоснабжения","выход","Решения: выходы",""),("Разводка канализации","компл","",""),
 ("Установка унитаза","шт","",""),("Установка раковины","шт","",""),("Установка смесителей","шт","",""),
 ("Установка ванны/душевой","компл","Решения: ванна/душ",""),("Установка полотенцесушителя","шт","Решения",""),
 ("Установка счётчиков воды","шт","Решения",""),("Краны, фильтр, редуктор","компл","",""),
 ("Отопление",None,None,None),
 ("Замена радиаторов + байпасы","шт","Решения: радиаторы",""),
 ("Электрика",None,None,None),
 ("Демонтаж старой проводки","компл","",""),("Штробление","м.п.","расчёт",""),
 ("Монтаж точек (розетки/выключатели/свет)","точка","Решения",""),("Прокладка кабеля","м","расчёт",""),
 ("Щит + автоматы + УЗО","компл","Решения",""),("Слаботочка","точка","Решения",""),
 ("Установка светильников","шт","Решения: свет",""),
 ("Двери",None,None,None),
 ("Установка межкомнатных дверей","шт","Замеры: проёмы",""),("Установка входной двери + откосы","шт","Замеры: проёмы",""),
 ("Доборы, наличники, фурнитура","компл","",""),
 ("Финал",None,None,None),
 ("Пусконаладка и проверка","компл","",""),("Финишная уборка","м²","Замеры: пол","")]
R2=[("Штукатурка стен по маякам (если кривизна)","м²","Решения: кривизна",""),
 ("Полная стяжка пола (если плита кривая)","м²","Решения: перепады",""),
 ("Демонтаж/переделка сантехкабины","компл","После вскрытия",""),
 ("Душевая вместо ванны (доплата)","компл","Решения: ванна/душ",""),
 ("Восстановление парапета + гидроизоляция плиты","компл","После вскрытия",""),
 ("Замена стояка (по проверке)","компл","Решения: стояк",""),
 ("Вывод и трасса под кондиционер","компл","Решения: кондиционер","")]
DECISIONS=[
 ("B. Полы","Граница плитки и ламината (входная зона)","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("B. Полы","Формат плитки на пол","РЕШЕНИЕ","Собственник","выбор","Сейчас","влияет на цену укладки и отход"),
 ("B. Полы","Перепады основания пола","ПОСЛЕ ВСКРЫТИЯ","Замер на объекте","мм","В процессе (после вскрытия)","наливной vs стяжка"),
 ("C. Стены/потолки","Чем покрыты стены и как держатся","ЗАМЕР","Замер на объекте","выбор","Домашнее задание","объём счистки"),
 ("C. Стены/потолки","Кривизна стен (после счистки)","ПОСЛЕ ВСКРЫТИЯ","Замер на объекте","мм","В процессе (после вскрытия)","шпаклёвка vs штукатурка"),
 ("C. Стены/потолки","Потолок — побелка мелом или краска","РЕШЕНИЕ","Замер на объекте","выбор","Сейчас","мел = смывка обязательна"),
 ("C. Стены/потолки","Галтель — ширина и материал","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("D. Санузел/кухня","Ванна или душевая","РЕШЕНИЕ","Собственник","выбор","Сейчас","финал по факту"),
 ("D. Санузел/кухня","Плитка в санузле — до потолка или до высоты","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("D. Санузел/кухня","Полотенцесушитель — водяной/электрический","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("D. Санузел/кухня","Стиралка — в санузле или на кухне","РЕШЕНИЕ","Собственник","выбор","Сейчас","вывод воды+слив+розетка"),
 ("D. Санузел/кухня","Кухонный гарнитур — мастер или мебельщики","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("E. Сантехника","Что реально топило соседей — где следы","ВНЕ КВАРТИРЫ","Соседи","выбор","Домашнее задание",""),
 ("E. Сантехника","Стояк: материал и состояние","ПОСЛЕ ВСКРЫТИЯ","Замер на объекте","выбор","В процессе (после вскрытия)",""),
 ("E. Сантехника","КСК: отключение/слив стояка, график отопсезона","ВНЕ КВАРТИРЫ","КСК/ОСИ","выбор","Домашнее задание",""),
 ("E. Сантехника","Соседи пустят ли для замены стояка","ВНЕ КВАРТИРЫ","Соседи","да/нет","Домашнее задание",""),
 ("E. Сантехника","Счётчики воды стоят? менять/ставить","СЧЁТ","Замер на объекте","шт","Домашнее задание",""),
 ("E. Сантехника","Сколько всего выходов воды","СЧЁТ","Собственник","выход","Сейчас",""),
 ("E. Сантехника","Сколько радиаторов меняем","СЧЁТ","Собственник","шт","Сейчас",""),
 ("E. Сантехника","Тип радиаторов","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("F. Электрика","Кухня: вся техника (под розетки + вывод на плиту)","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("F. Электрика","Розетки по комнатам — сколько и где","РЕШЕНИЕ","Собственник","шт","Сейчас",""),
 ("F. Электрика","Свет: люстры/точечные/бра — сколько","РЕШЕНИЕ","Собственник","шт","Сейчас",""),
 ("F. Электрика","Кондиционер — будет? где","РЕШЕНИЕ","Собственник","да/нет","Сейчас","трасса до штробления"),
 ("F. Электрика","Слаботочка: интернет/ТВ/домофон","РЕШЕНИЕ","Собственник","точка","Сейчас",""),
 ("F. Электрика","Заземление (3 провода) или двухпроводка","ЗАМЕР","Замер на объекте","выбор","Домашнее задание",""),
 ("F. Электрика","Щит — где, менять/ревизия","ЗАМЕР","Замер на объекте","выбор","Домашнее задание",""),
 ("G. Двери/балкон","Тип/материал дверей, со стеклом","РЕШЕНИЕ","Собственник","выбор","Сейчас","размеры — на листе замеров"),
 ("G. Двери/балкон","Окна (не балкон) — менять или оставить","РЕШЕНИЕ","Собственник","да/нет","Сейчас",""),
 ("G. Двери/балкон","Балкон: остекление, утепление, отделка","РЕШЕНИЕ","Собственник","выбор","Домашнее задание",""),
 ("G. Двери/балкон","Балконная плита/парапет — состояние","ПОСЛЕ ВСКРЫТИЯ","Замер на объекте","выбор","В процессе (после вскрытия)",""),
 ("H. Материалы","Класс по группам (эконом/комфорт)","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("H. Материалы","Кто и когда покупает; простой при опоздании","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("I. Труд/организация","Мастер делает всё сам или часть отдаёт","СПРОСИТЬ МАСТЕРА","Мастер","выбор","Сейчас","цены субподряда отдельно"),
 ("I. Труд/организация","Ставка мастера (день/месяц), сколько человек","СПРОСИТЬ МАСТЕРА","Мастер","деньги","Сейчас","без этого не проверить сумму"),
 ("I. Труд/организация","Сколько дней/недель на весь объём","СПРОСИТЬ МАСТЕРА","Мастер","выбор","Сейчас",""),
 ("I. Труд/организация","Квартира пустая/жилая; дедлайн","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("J. Договор/деньги","График оплаты (предоплата/этапы/удержание)","РЕШЕНИЕ","Собственник","выбор","Сейчас",""),
 ("J. Договор/деньги","Гарантия мастера — на что и сколько","РЕШЕНИЕ","Мастер","выбор","Сейчас",""),
 ("J. Договор/деньги","Приёмка по этапам + фотофиксация скрытых работ","РЕШЕНИЕ","Собственник","да/нет","Сейчас","")]

# --- Замеры→Смета: однозначные агрегатные привязки (Fix 3) ---
# Кол-во(D) проставляется ФОРМУЛОЙ на ИТОГО листа «Замеры — стены и пол».
# Иначе (поштучно/компл/точка/рейс/осмотр/неоднозначно) — ручной ввод.
# "wall" → ИТОГО стены нетто (I), "perim" → ИТОГО периметр (F), "floor" → ИТОГО пол (J).
AGG_BY_SRC={"Замеры: стены нетто":"wall","Замеры: периметр":"perim"}
# Позиции «по всему полу/потолку» — итог пола. Только эти имена (не выдумываем связи).
AGG_FLOOR_NAMES={
 "Грунтовка основания",
 "Наливной пол / ровнитель",
 "Смывка побелки",
 "Шпаклёвка потолков + шлифовка",
 "Грунт + покраска потолков (2 слоя)",
}


def build(cfg: dict, out_path: str) -> dict:
    city=cfg.get("city","").strip()
    cur=cfg.get("currency","₸").strip() or "₸"
    title=cfg.get("title","").strip()
    rooms=cfg.get("rooms") or DEFAULT_ROOMS
    openings=cfg.get("openings") or DEFAULT_OPENINGS
    materials=[(m["name"],m.get("unit",""),m.get("waste")) for m in cfg["materials"]] if cfg.get("materials") else DEFAULT_MATERIALS
    MONEY=f'#,##0" {cur}"'
    head_city=f" · {city}" if city else ""
    head_title=f" · {title}" if title else ""
    # Геометрия строк листа «Замеры — стены и пол» (данные с 4-й строки, ИТОГО ниже).
    Wn="Замеры — стены и пол"; Pn="Замеры — проёмы"
    wf=4; wl_last=wf+len(rooms)-1; wl_total=wl_last+1
    # «Замеры — проёмы»: реальные проёмы (без хвостовых строк «(добавить)»).
    op_first=4; op_data_last=op_first+len(openings)-1
    # B2: параметры расчёта мусора живут на «Сводке» (фикс. ячейки B9/B10).
    # Рейсы = ROUNDUP(площадь пола × коэф (м³/м²) / вместимость рейса (м³)).
    SV_COEF="'Сводка'!$B$9"; SV_TRUCK="'Сводка'!$B$10"
    DEBRIS=(f'=IF(OR(N(\'{Wn}\'!J{wl_total})=0,N({SV_TRUCK})=0),"",'
            f'ROUNDUP(N(\'{Wn}\'!J{wl_total})*N({SV_COEF})/N({SV_TRUCK}),0))')

    # Предрасчёт геометрии листа «Смета» (нужен тестам, которые строятся раньше Сметы).
    # Возвращает также карту {row: kind} привязок Кол-во к ИТОГО листа стен/пола.
    def _agg_kind(name,src):
        if src in AGG_BY_SRC: return AGG_BY_SRC[src]
        if name in AGG_FLOOR_NAMES: return "floor"
        return None
    def _block_geom(items,start,idx,bind):
        r=start+1; fw=None  # +1 — строка-заголовок раздела
        for it in items:
            if len(it)==4 and it[1] is None: r+=1; continue  # подзаголовок
            name,unit,src,note=it; fw=fw or r
            k=_agg_kind(name,src)
            if k: bind[r]=k
            idx+=1; r+=1
        return r,fw,r-1,idx
    SM_BIND={}  # {Смета-строка: "wall"/"perim"/"floor"}
    _r=4; _r,_f1,_l1,_n=_block_geom(R1,_r,1,SM_BIND)   # _f1 — первая позиция Р1
    _r+=2; _r,_f2,_l2,_n=_block_geom(R2,_r,_n,SM_BIND)  # _l2 — последняя позиция Р2
    SMn="Смета"
    SM_PRICE_ALL=f"'{SMn}'!$E${_f1}:$E${_l2}"   # цены Р1+Р2 (для гейта источника)
    SM_SRC_ALL=f"'{SMn}'!$G${_f1}:$G${_l2}"     # «Ценовой источник (URL)» Р1+Р2
    SM_WALL_ROWS=[r for r,k in SM_BIND.items() if k=="wall"]  # строки, привязанные к ИТОГО нетто

    def ghost(ws,r,c,hint,numfmt='#,##0',h="center"):
        SC(ws.cell(row=r,column=c,value=hint),F(9,it=True,color=GHOSTC),INPUT,AL(h),numfmt=numfmt)
    def inp(ws,r,c,v,numfmt='#,##0',h="center"):
        SC(ws.cell(row=r,column=c,value=v),F(10,color="0000FF"),INPUT,AL(h),numfmt=numfmt)
    def calc(ws,r,c,formula,numfmt='#,##0.00',bold=False,fillc=CALC):
        SC(ws.cell(row=r,column=c,value=formula),F(10,b=bold),fillc,AL("center"),numfmt=numfmt)
    def title_rows(ws,ncol,t,sub):
        last=chr(64+ncol); ws.merge_cells(f"A1:{last}1"); SC(ws["A1"],F(13,True,WHITE),NAVY,AL("center"),bd=False)
        ws["A1"].value=t; ws.row_dimensions[1].height=32
        ws.merge_cells(f"A2:{last}2"); SC(ws["A2"],F(9,False,"444444"),GREY,AL("left"),bd=False)
        ws["A2"].value=sub; ws.row_dimensions[2].height=30
    def header(ws,headers,row=3):
        for i,h in enumerate(headers,start=1): SC(ws.cell(row=row,column=i,value=h),F(10,True,WHITE),HEAD,AL("center"))
        ws.row_dimensions[row].height=26

    wb=Workbook()
    # Справочники
    sp=wb.active; sp.title="Справочники"; sp.sheet_properties.tabColor="808080"
    rooms_names=[r["name"] for r in rooms]
    el_names=sorted({o["element"] for o in openings}) or ["Дверь","Окно"]
    lists={"A":("Ед. изм.",["м²","м.п.","м","мм","м³","шт","компл","точка","выход","да/нет","выбор","деньги","%","рейс"]),
           "B":("Тип",["РЕШЕНИЕ","ЗАМЕР","СЧЁТ","ВНЕ КВАРТИРЫ","ПОСЛЕ ВСКРЫТИЯ","СПРОСИТЬ МАСТЕРА"]),
           "C":("Кто отвечает",["Собственник","Мастер","КСК/ОСИ","Соседи","Замер на объекте"]),
           "D":("Когда",["Сейчас","Домашнее задание","В процессе (после вскрытия)"]),
           "F":("Помещение",rooms_names),"G":("Элемент",el_names)}
    maxrow={}
    for col,(t,vals) in lists.items():
        SC(sp[f"{col}1"],F(10,True,WHITE),HEAD,AL("center")); sp[f"{col}1"].value=t
        for i,v in enumerate(vals,start=2): SC(sp[f"{col}{i}"],F(10),None,AL("left")); sp[f"{col}{i}"].value=v
        sp.column_dimensions[col].width=24; maxrow[col]=1+len(vals)
    def rng(col): return f"'Справочники'!${col}$2:${col}${maxrow[col]}"

    # Замеры — проёмы
    pr=wb.create_sheet("Замеры — проёмы"); pr.sheet_properties.tabColor="C55A11"
    header(pr,["№","Элемент","Помещение","Ширина, мм","Высота, мм","Кол-во, шт","Площадь, м²","Примечание","Высота комнаты, мм","Превышение"])
    for i,w in enumerate([5,24,14,13,13,10,12,26,15,12],start=1): pr.column_dimensions[chr(64+i)].width=w
    title_rows(pr,10,f"ЗАМЕРЫ ПРОЁМОВ{head_title}{head_city}","Строка = один проём. Ширина/Высота — в мм поверх подсказки. «Площадь» считается сама и вычитается из стен по «Помещение». Столбцы I/J — служебные (высота комнаты и флаг превышения для тестов).")
    # Высота комнаты (I) = SUMIF по листу стен: Помещение = строка проёма → высота. Флаг (J) = проём выше комнаты.
    WLB=f"'{Wn}'!$B${wf}:$B${wl_last}"; WLE=f"'{Wn}'!$E${wf}:$E${wl_last}"
    r=4; rows_op=openings+[{"element":"(добавить)","room":"","qty":None},{"element":"(добавить)","room":"","qty":None}]
    for i,o in enumerate(rows_op,start=1):
        SC(pr.cell(row=r,column=1,value=i),F(9),None,AL("center"))
        SC(pr.cell(row=r,column=2,value=o.get("element","")),F(10),None,AL("left"))
        SC(pr.cell(row=r,column=3,value=o.get("room","")),F(10),None,AL("left"))
        ghost(pr,r,4,"ширина, мм"); ghost(pr,r,5,"высота, мм")
        if o.get("qty"): inp(pr,r,6,o["qty"])
        else: ghost(pr,r,6,"кол-во")
        calc(pr,r,7,f"=N(D{r})*N(E{r})*N(F{r})/1000000",numfmt='#,##0.00',fillc=CALC)
        SC(pr.cell(row=r,column=8,value=o.get("note","")),F(9,False,"555555"),None,AL("left"))
        calc(pr,r,9,f"=SUMIF({WLB},C{r},{WLE})",numfmt='#,##0',fillc=CALC)
        calc(pr,r,10,f"=IF(AND(ISNUMBER(E{r}),I{r}>0,E{r}>I{r}),1,0)",numfmt='#,##0',fillc=CALC)
        pr.row_dimensions[r].height=20; r+=1
    pr_last=r-1
    pr.freeze_panes="A4"; pr.auto_filter.ref=f"A3:J{pr_last}"
    dv=DataValidation(type="list",formula1=rng("G"),allow_blank=True); pr.add_data_validation(dv); dv.add(f"B4:B{pr_last}")
    dv=DataValidation(type="list",formula1=rng("F"),allow_blank=True); pr.add_data_validation(dv); dv.add(f"C4:C{pr_last}")
    im=DataValidation(showInputMessage=True,showErrorMessage=False); im.promptTitle="Замер"; im.prompt="Введите размер в миллиметрах"
    pr.add_data_validation(im); im.add(f"D4:E{pr_last}")

    # Замеры — стены и пол
    wl=wb.create_sheet("Замеры — стены и пол"); wl.sheet_properties.tabColor="C55A11"
    header(wl,["№","Помещение","Длина, мм","Ширина, мм","Высота, мм","Периметр, м","Стены брутто, м²","Проёмы, м²","Стены нетто, м²","Пол, м²","Примечание"])
    for i,w in enumerate([4,13,12,12,12,11,13,11,13,10,24],start=1): wl.column_dimensions[chr(64+i)].width=w
    title_rows(wl,11,f"ЗАМЕРЫ СТЕН И ПОЛА{head_title}{head_city}","Брутто = периметр × высота. Проёмы — авто-сумма из листа «Замеры — проёмы». Нетто = брутто − проёмы (под штукатурку/покраску). Пол — из тех.паспорта; высоту перемерить.")
    PRC=f"'Замеры — проёмы'!$C$4:$C${pr_last}"; PRG=f"'Замеры — проёмы'!$G$4:$G${pr_last}"
    r=4; wf=r
    for i,rm in enumerate(rooms,start=1):
        SC(wl.cell(row=r,column=1,value=i),F(9),None,AL("center"))
        SC(wl.cell(row=r,column=2,value=rm["name"]),F(10,True),None,AL("left"))
        ghost(wl,r,3,"длина, мм"); ghost(wl,r,4,"ширина, мм"); ghost(wl,r,5,"высота, мм")
        calc(wl,r,6,f"=2*(N(C{r})+N(D{r}))/1000",numfmt='#,##0.00')
        calc(wl,r,7,f"=F{r}*N(E{r})/1000",numfmt='#,##0.0',bold=True)
        calc(wl,r,8,f'=SUMIF({PRC},B{r},{PRG})',numfmt='#,##0.0')
        calc(wl,r,9,f"=MAX(G{r}-H{r},0)",numfmt='#,##0.0',bold=True)
        if rm.get("floor"): inp(wl,r,10,rm["floor"],numfmt='#,##0.0')
        else: ghost(wl,r,10,"пол, м²",numfmt='#,##0.0')
        SC(wl.cell(row=r,column=11,value=rm.get("note","")),F(9,False,"555555"),None,AL("left"))
        wl.row_dimensions[r].height=20; r+=1
    wl_last=r-1
    # Итог периметра (F) нужен для привязки Кол-ва в Смете (Fix 3); раньше F был под merge.
    wl.merge_cells(f"A{r}:E{r}"); SC(wl[f"A{r}"],F(10,True),TOTAL,AL("right")); wl[f"A{r}"].value="ИТОГО"
    for col,L in [(6,"F"),(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
        SC(wl.cell(row=r,column=col,value=f"=SUM({L}{wf}:{L}{wl_last})"),F(10,True),TOTAL,AL("center"),numfmt='#,##0.0')
    SC(wl.cell(row=r,column=11),F(9),TOTAL); wl.freeze_panes="A4"
    im=DataValidation(showInputMessage=True,showErrorMessage=False); im.promptTitle="Замер"; im.prompt="Введите размер в миллиметрах"
    wl.add_data_validation(im); im.add(f"C4:E{wl_last}")

    # Тесты замеров
    ts=wb.create_sheet("Тесты замеров"); ts.sheet_properties.tabColor="A6300F"
    for i,w in enumerate([4,40,18,12,9,26,34,9],start=1): ts.column_dimensions[chr(64+i)].width=w
    ts.merge_cells("A1:H1"); SC(ts["A1"],F(13,True,WHITE),NAVY,AL("center"),bd=False)
    ts["A1"].value=f"ТЕСТЫ ЗАМЕРОВ (контроль качества){head_title}{head_city}"; ts.row_dimensions[1].height=32
    ts.merge_cells("A2:H2"); SC(ts["A2"],F(9,False,"444444"),GREY,AL("left"),bd=False)
    ts["A2"].value="Вставь замеры — тесты пересчитаются. ✗ = ошибка, ⚠ = проверь, OK = норма, — = нет данных. Столбец «Уровень»: 2=✗, 1=⚠, 0=ок (для портируемой сводки)."; ts.row_dimensions[2].height=26
    ts.merge_cells("A3:H3"); SC(ts["A3"],F(12,True,"1F3864"),"FFF2CC",AL("center")); ts.row_dimensions[3].height=26
    header(ts,["№","Проверка","Норма / ожидание","Факт","Откл.","Статус","Что делать","Уровень"],row=5)
    rr=[6]
    def banner(t):
        r=rr[0]; ts.merge_cells(f"A{r}:H{r}"); SC(ts[f"A{r}"],F(11,True,"1F3864"),SECT,AL("left")); ts[f"A{r}"].value=t
        ts.row_dimensions[r].height=20; rr[0]+=1
    def add(name,norma,fact,otkl,status,action,fnf='#,##0.00'):
        r=rr[0]
        SC(ts.cell(row=r,column=1,value=r-5),F(9),None,AL("center"))
        SC(ts.cell(row=r,column=2,value=name),F(10),None,AL("left"))
        SC(ts.cell(row=r,column=3,value=norma),F(9),None,AL("center"))
        SC(ts.cell(row=r,column=4,value=fact),F(10),CALC,AL("center"),numfmt=fnf)
        SC(ts.cell(row=r,column=5,value=otkl if otkl else ""),F(9),CALC,AL("center"),numfmt='0.0%')
        SC(ts.cell(row=r,column=6,value=status),F(10,True),None,AL("center"))
        SC(ts.cell(row=r,column=7,value=action),F(9,False,"555555"),None,AL("left"))
        # «Уровень» (Fix 6): числовой код статуса для портируемой сводки (без wildcard-COUNTIF).
        SC(ts.cell(row=r,column=8,value=f'=IF(ISNUMBER(SEARCH("✗",F{r})),2,IF(ISNUMBER(SEARCH("⚠",F{r})),1,0))'),F(9),CALC,AL("center"),numfmt='0')
        ts.row_dimensions[r].height=22; rr[0]+=1
    banner("Площадь пола: обмер (Д×Ш) против тех.паспорта")
    for idx,rm in enumerate(rooms):
        wr=4+idx; g=f"AND(ISNUMBER('{Wn}'!C{wr}),ISNUMBER('{Wn}'!D{wr}),ISNUMBER('{Wn}'!J{wr}))"
        a=f"'{Wn}'!C{wr}*'{Wn}'!D{wr}/1000000"; e=f"'{Wn}'!J{wr}"
        add(f"Пол «{rm['name']}» = паспорт", f"='{Wn}'!J{wr}",
            f'=IF({g},{a},"—")', f'=IF({g},({a}-{e})/{e},"—")',
            f'=IF(NOT({g}),"— нет данных",IF(ABS(({a}-{e})/{e})<=0.10,"OK","⚠ проверь Д/Ш или паспорт"))',
            "обмер и паспорт должны сходиться (±10%)", fnf='#,##0.0')
    banner("Высоты потолков: согласованность и единицы")
    add("Разброс высот по комнатам","≤ 50 мм",
        f'=IF(COUNT(\'{Wn}\'!E{wf}:E{wl_last})<2,"—",MAX(\'{Wn}\'!E{wf}:E{wl_last})-MIN(\'{Wn}\'!E{wf}:E{wl_last}))',None,
        f'=IF(COUNT(\'{Wn}\'!E{wf}:E{wl_last})<2,"— нет данных",IF(MAX(\'{Wn}\'!E{wf}:E{wl_last})-MIN(\'{Wn}\'!E{wf}:E{wl_last})<=50,"OK",IF(MAX(\'{Wn}\'!E{wf}:E{wl_last})-MIN(\'{Wn}\'!E{wf}:E{wl_last})<=150,"⚠ небольшой разброс","✗ высоты не сходятся")))',
        "у одной плиты высота одинаковая", fnf='#,##0')
    add("Средняя высота правдоподобна (2000–3500 мм)","2000–3500 мм",
        f'=IF(COUNT(\'{Wn}\'!E{wf}:E{wl_last})=0,"—",ROUND(AVERAGE(\'{Wn}\'!E{wf}:E{wl_last}),0))',None,
        f'=IF(COUNT(\'{Wn}\'!E{wf}:E{wl_last})=0,"— нет данных",IF(AND(AVERAGE(\'{Wn}\'!E{wf}:E{wl_last})>=2000,AVERAGE(\'{Wn}\'!E{wf}:E{wl_last})<=3500),"OK","✗ похоже не в мм"))',
        "иначе введено в см/м, а нужно мм", fnf='#,##0')
    banner("Проёмы: непротиворечивость со стенами")
    add("Проёмы в комнате не больше площади стен","0 нарушений",
        f"=SUMPRODUCT(('{Wn}'!G{wf}:G{wl_last}>0)*('{Wn}'!H{wf}:H{wl_last}>='{Wn}'!G{wf}:G{wl_last}))",None,
        f'=IF(SUMPRODUCT(--(\'{Wn}\'!G{wf}:G{wl_last}>0))=0,"— нет данных",IF(SUMPRODUCT((\'{Wn}\'!G{wf}:G{wl_last}>0)*(\'{Wn}\'!H{wf}:H{wl_last}>=\'{Wn}\'!G{wf}:G{wl_last}))=0,"OK","✗ проёмы ≥ стены"))',
        "проём не может быть больше стены", fnf='#,##0')
    # Построчно (Fix 4): флаг J на «Замеры — проёмы» = высота проёма > высоты своей комнаты.
    add("Высота проёма ≤ высоты комнаты (построчно)","0 превышений",
        f"=SUM('{Pn}'!J{op_first}:J{op_data_last})",None,
        f'=IF(COUNT(\'{Pn}\'!E{op_first}:E{op_data_last})=0,"— нет данных",IF(SUM(\'{Pn}\'!J{op_first}:J{op_data_last})>0,"✗ проём выше комнаты","OK"))',
        "дверь/окно выше потолка своей комнаты = опечатка", fnf='#,##0')
    banner("Единицы и правдоподобие (ловит см/м вместо мм)")
    add("Размеры проёмов в 300–3500 мм","0 вне нормы",
        f"=SUMPRODUCT(ISNUMBER('{Pn}'!D{op_first}:D{op_data_last})*(('{Pn}'!D{op_first}:D{op_data_last}<300)+('{Pn}'!D{op_first}:D{op_data_last}>3500)))+SUMPRODUCT(ISNUMBER('{Pn}'!E{op_first}:E{op_data_last})*(('{Pn}'!E{op_first}:E{op_data_last}<300)+('{Pn}'!E{op_first}:E{op_data_last}>3500)))",None,
        f'=IF(SUMPRODUCT(ISNUMBER(\'{Pn}\'!D{op_first}:D{op_data_last})*((\'{Pn}\'!D{op_first}:D{op_data_last}<300)+(\'{Pn}\'!D{op_first}:D{op_data_last}>3500)))+SUMPRODUCT(ISNUMBER(\'{Pn}\'!E{op_first}:E{op_data_last})*((\'{Pn}\'!E{op_first}:E{op_data_last}<300)+(\'{Pn}\'!E{op_first}:E{op_data_last}>3500)))=0,"OK","⚠ размеры вне нормы (см/м?)")',
        "проём обычно 300–3500 мм", fnf='#,##0')
    banner("Полнота (data-range без хвостовых строк «(добавить)»)")
    # Мультипликативная форма (без --NOT-над-массивом) — переносима в Numbers (Fix 4/6).
    INC=(f"SUMPRODUCT((NOT(ISNUMBER('{Wn}'!C{wf}:E{wl_last})))*1)"
         f"+SUMPRODUCT((NOT(ISNUMBER('{Pn}'!D{op_first}:E{op_data_last})))*1)")
    add("Незаполненные ячейки замеров","0",
        f"={INC}",None,
        f'=IF({INC}=0,"OK — все замеры внесены","— остались незаполненные")',
        "пока не 0 — обмер не закончен", fnf='#,##0')
    banner("Смета: провенанс цен и привязка кол-ва")
    # Гейт источника цены (Fix 2): цена заполнена (ISNUMBER), но «Ценовой источник (URL)» пуст.
    # Мультипликативная форма SUMPRODUCT (как в остальных тестах) — переносима в Numbers.
    GATE=f'SUMPRODUCT(ISNUMBER({SM_PRICE_ALL})*({SM_SRC_ALL}=""))'
    add("Цена без ценового источника","0",
        f"={GATE}",None,
        f'=IF({GATE}=0,"OK — у каждой цены есть источник","✗ есть цена без источника")',
        "цена без URL/оффера — непроверяемо", fnf='#,##0')
    # Реконсиляция (Fix 3): авто-привязанные по стенам Кол-во должны равняться ИТОГО нетто.
    if SM_WALL_ROWS:
        neq="+".join([f"--('{SMn}'!D{rw}<>'{Wn}'!I{wl_total})" for rw in SM_WALL_ROWS])
        sumD="+".join([f"N('{SMn}'!D{rw})" for rw in SM_WALL_ROWS])
        add("Кол-во «по стенам» = ИТОГО нетто","0 расхождений",
            f"=({sumD})-{len(SM_WALL_ROWS)}*'{Wn}'!I{wl_total}",None,
            f'=IF(SUMPRODUCT({neq})=0,"OK — привязка к ИТОГО нетто цела","✗ привязка кол-ва разошлась")',
            "каждая строка «по стенам» берёт ИТОГО нетто", fnf='#,##0.0')
    tlast=rr[0]-1
    # Баннер-сводка (Fix 6): портируемо — SUMPRODUCT по числовому «Уровень», без wildcard COUNTIF.
    # Мультипликативная форма (=2)*1 эквивалентна --(=2), но переносима в Numbers и однотипна с тестами.
    ts["A3"].value=f'="✗ ошибок: "&SUMPRODUCT(($H$6:$H${tlast}=2)*1)&"      ⚠ проверить: "&SUMPRODUCT(($H$6:$H${tlast}=1)*1)'
    ts.freeze_panes="A6"; ts.auto_filter.ref=f"A5:H{tlast}"
    cr=f"F6:F{tlast}"
    ts.conditional_formatting.add(cr,FormulaRule(formula=['ISNUMBER(SEARCH("✗",F6))'],fill=FILL("FFC7CE"),font=Font(name=AR,bold=True,color="9C0006")))
    ts.conditional_formatting.add(cr,FormulaRule(formula=['ISNUMBER(SEARCH("⚠",F6))'],fill=FILL("FFEB9C"),font=Font(name=AR,color="9C6500")))
    ts.conditional_formatting.add(cr,FormulaRule(formula=['ISNUMBER(SEARCH("OK",F6))'],fill=FILL("C6EFCE"),font=Font(name=AR,color="006100")))
    ts.conditional_formatting.add(cr,FormulaRule(formula=['LEFT(F6,1)="—"'],fill=FILL("E7E6E6"),font=Font(name=AR,color="808080")))

    # Решения и проверки
    op=wb.create_sheet("Решения и проверки"); op.sheet_properties.tabColor=HEAD
    for i,w in enumerate([5,22,50,16,17,22,9,24,26],start=1): op.column_dimensions[chr(64+i)].width=w
    title_rows(op,9,f"РЕШЕНИЯ И ПРОВЕРКИ{head_title}{head_city}","Это НЕ замеры. Впиши ответ поверх подсказки. «Когда»: Сейчас / Домашнее задание / После вскрытия.")
    header(op,["№","Раздел","Вопрос","Тип","Кто отвечает","Ответ","Ед. изм.","Когда","Комментарий"])
    SECCOLOR={"B":"DEEBF7","C":"FFF2CC","D":"FCE4D6","E":"E2EFDA","F":"DEEBF7","G":"FFF2CC","H":"FCE4D6","I":"E2EFDA","J":"DEEBF7"}
    HINT={"да/нет":"да / нет","выбор":"выбери / впиши","шт":"кол-во, шт","точка":"кол-во","выход":"кол-во","деньги":f"сумма, {cur}","мм":"мм"}
    r=4
    for i,(sec,text,typ,who,unit,when,com) in enumerate(DECISIONS,start=1):
        rf=SECCOLOR.get(sec.split(".")[0],"FFFFFF")
        SC(op.cell(row=r,column=1,value=i),F(10,True),rf,AL("center"))
        SC(op.cell(row=r,column=2,value=sec),F(9),rf,AL("left"))
        SC(op.cell(row=r,column=3,value=text),F(10),rf,AL("left"))
        SC(op.cell(row=r,column=4,value=typ),F(9),rf,AL("center"))
        SC(op.cell(row=r,column=5,value=who),F(9),rf,AL("center"))
        ghost(op,r,6,HINT.get(unit,"впиши ответ"),numfmt=None,h="left")
        SC(op.cell(row=r,column=7,value=unit),F(9),rf,AL("center"))
        SC(op.cell(row=r,column=8,value=when),F(9),rf,AL("center"))
        SC(op.cell(row=r,column=9,value=com),F(9),rf,AL("left"))
        op.row_dimensions[r].height=28; r+=1
    op_last=r-1; op.freeze_panes="A4"; op.auto_filter.ref=f"A3:I{op_last}"
    for col,src in [("D","B"),("E","C"),("G","A"),("H","D")]:
        d=DataValidation(type="list",formula1=rng(src),allow_blank=True); op.add_data_validation(d); d.add(f"{col}4:{col}{op_last}")

    # Смета
    sm=wb.create_sheet("Смета"); sm.sheet_properties.tabColor="548235"
    for i,w in enumerate([5,50,9,10,14,16,26,13,15,20,20],start=1): sm.column_dimensions[chr(64+i)].width=w
    title_rows(sm,11,f"СМЕТА НА РЕМОНТ (работа; материалы отдельно){head_title}{head_city}",f"Кол-во — формулой с «Замеры — стены и пол» (стены→нетто, пол→ИТОГО пола, периметр→ИТОГО периметра), иначе ручной ввод. Цена ({cur}) — из живых офферов; рядом обязателен ценовой источник (URL), дата и тип.")
    header(sm,["№","Вид работ","Ед. изм.","Кол-во",f"Цена за ед., {cur}",f"Сумма, {cur}","Ценовой источник (URL)","Дата котировки","Тип","Источник данных","Примечание"])
    # Куда смотрит формула Кол-во для каждого вида привязки (ИТОГО листа «Замеры — стены и пол»).
    AGG_CELL={"wall":f"'{Wn}'!I{wl_total}","perim":f"'{Wn}'!F{wl_total}","floor":f"'{Wn}'!J{wl_total}"}
    def block(title,items,start,idx,r2=False):
        r=start; sm.merge_cells(f"A{r}:K{r}"); SC(sm[f"A{r}"],F(12,True,WHITE if r2 else "1F3864"),"C55A11" if r2 else SECT,AL("left"))
        sm[f"A{r}"].value=title; sm.row_dimensions[r].height=24; r+=1; fw=None
        for it in items:
            if len(it)==4 and it[1] is None:
                sm.merge_cells(f"A{r}:K{r}"); SC(sm[f"A{r}"],F(10,True,"1F3864"),SUB,AL("left")); sm[f"A{r}"].value="   "+it[0]
                sm.row_dimensions[r].height=18; r+=1; continue
            name,unit,src,note=it; fw=fw or r; rf=R2FILL if r2 else None
            SC(sm.cell(row=r,column=1,value=idx),F(9),rf,AL("center")); idx+=1
            SC(sm.cell(row=r,column=2,value=name),F(10),rf,AL("left"))
            SC(sm.cell(row=r,column=3,value=unit),F(9),rf,AL("center"))
            kind=_agg_kind(name,src)
            if name=="Вынос и вывоз мусора":  # B2: рейсы = объём демонтажа ÷ вместимость (а не «с потолка»)
                calc(sm,r,4,DEBRIS,numfmt='#,##0',bold=False); data_src="мусор: пол×коэф÷рейс"
            elif kind:  # Кол-во — формулой на ИТОГО листа замеров (Fix 3)
                calc(sm,r,4,f"={AGG_CELL[kind]}",numfmt='#,##0.00',bold=False)
                data_src=src  # сохраняем исходную привязку («Замеры: …»)
            else:     # неоднозначно/поштучно — ручной ввод
                ghost(sm,r,4,"кол-во"); data_src="ручной ввод"
            ghost(sm,r,5,f"цена, {cur}")
            SC(sm.cell(row=r,column=6,value=f"=N(D{r})*N(E{r})"),F(10,True),rf,AL("center"),numfmt=MONEY)
            # «Ценовой источник (URL)» — пустая input-ячейка (без подсказки-текста),
            # чтобы гейт «цена есть, источник пуст» (G="") работал корректно (Fix 2).
            SC(sm.cell(row=r,column=7),F(10,color="0000FF"),INPUT,AL("left"),numfmt=None)
            ghost(sm,r,8,"дд.мм.гггг",numfmt=None)                # Дата котировки
            ghost(sm,r,9,"оффер/…",numfmt=None)                   # Тип (дропдаун ниже)
            SC(sm.cell(row=r,column=10,value=data_src),F(9),rf,AL("center"))  # Источник данных
            SC(sm.cell(row=r,column=11,value=note),F(9,False,"555555"),rf,AL("left"))
            sm.row_dimensions[r].height=20; r+=1
        return r,fw,r-1,idx
    def total_row(label,formula,fillc,white=False,fsz=11):
        sm.merge_cells(f"A{row}:E{row}"); SC(sm[f"A{row}"],F(fsz,True,WHITE if white else "000000"),fillc,AL("right")); sm[f"A{row}"].value=label
        SC(sm.cell(row=row,column=6,value=formula),F(fsz,True,WHITE if white else "000000"),fillc,AL("center"),numfmt=MONEY)
        for c in range(7,12): SC(sm.cell(row=row,column=c),F(9,False,WHITE if white else "000000"),fillc)
    row=4; row,f1,l1,nidx=block("РАЗДЕЛ 1 — ДЕЛАЕМ ТОЧНО",R1,row,1,False)
    P1=f"E{f1}:E{l1}"  # диапазон цен Раздела 1
    # ИТОГ-0 (Fix 5): при пустых ценах — «— цены не внесены», а не 0.
    total_row("ИТОГО РАЗДЕЛ 1 (работа)",f'=IF(COUNT({P1})=0,"— цены не внесены",SUM(F{f1}:F{l1}))',TOTAL)
    row+=2; row,f2,l2,nidx=block("РАЗДЕЛ 2 — МОЖЕТ ПОНАДОБИТЬСЯ",R2,row,nidx,True)
    total_row("ИТОГО РАЗДЕЛ 2 (условные)",f"=SUM(F{f2}:F{l2})",R2FILL)
    t2=row; row+=2
    sm.merge_cells(f"A{row}:D{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Резерв на непредвиденное, % от Раздела 1"
    inp(sm,row,5,0.15,numfmt='0%')
    SC(sm.cell(row=row,column=6,value=f'=IF(COUNT({P1})=0,"— цены не внесены",SUM(F{f1}:F{l1})*N(E{row}))'),F(10,True),GREY,AL("center"),numfmt=MONEY)
    for c in range(7,12): SC(sm.cell(row=row,column=c),F(9),GREY)
    rres=row; row+=1
    total_row("ВСЕГО (Раздел 1 + резерв)",f'=IF(COUNT({P1})=0,"— цены не внесены",SUM(F{f1}:F{l1})*(1+N(E{rres})))',"1F3864",white=True,fsz=12)
    # B1: сверка труда — вторая независимая оценка (бригада×срок×ставка) против сметы по позициям.
    row+=2
    sm.merge_cells(f"A{row}:K{row}"); SC(sm[f"A{row}"],F(11,True,"1F3864"),SECT,AL("left"))
    sm[f"A{row}"].value="СВЕРКА ТРУДА (вторая оценка снизу-сверху; в ВСЕГО не входит)"; sm.row_dimensions[row].height=22; row+=1
    sm.merge_cells(f"A{row}:D{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value=f"Ставка мастера за день, {cur}"
    ghost(sm,row,5,"ставка/день",numfmt=MONEY); _rate=row; row+=1
    sm.merge_cells(f"A{row}:D{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Человек в бригаде"
    ghost(sm,row,5,"чел",numfmt='#,##0'); _ppl=row; row+=1
    sm.merge_cells(f"A{row}:D{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Рабочих дней"
    ghost(sm,row,5,"дней",numfmt='#,##0'); _days=row; row+=1
    _fund=f"N(E{_rate})*N(E{_ppl})*N(E{_days})"; _pos=f"SUM(F{f1}:F{l1})"
    sm.merge_cells(f"A{row}:E{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Фонд по бригаде = ставка×чел×дней"
    SC(sm.cell(row=row,column=6,value=f"={_fund}"),F(10,True),GREY,AL("center"),numfmt=MONEY)
    for _c in range(7,12): SC(sm.cell(row=row,column=_c),F(9),GREY)
    row+=1
    sm.merge_cells(f"A{row}:E{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Сумма позиций Раздела 1 (для сверки)"
    SC(sm.cell(row=row,column=6,value=f'=IF(COUNT({P1})=0,"— цены не внесены",{_pos})'),F(10,True),GREY,AL("center"),numfmt=MONEY)
    for _c in range(7,12): SC(sm.cell(row=row,column=_c),F(9),GREY)
    row+=1
    sm.merge_cells(f"A{row}:E{row}"); SC(sm[f"A{row}"],F(10,True),GREY,AL("right")); sm[f"A{row}"].value="Статус сверки труда"
    SC(sm.cell(row=row,column=6,value=f'=IF(OR(COUNT({P1})=0,{_fund}=0,{_pos}=0),"— нет данных",IF(ABS({_fund}-{_pos})/{_pos}<=0.25,"OK: порядок труда сходится (±25%)","⚠ расходится — проверь ставку/дни или цены"))'),F(10,True),GREY,AL("center"))
    for _c in range(7,12): SC(sm.cell(row=row,column=_c),F(9),GREY)
    row+=1
    sm.merge_cells(f"A{row}:K{row}"); SC(sm[f"A{row}"],F(9,False,"555555"),GREY,AL("left"),bd=False)
    sm[f"A{row}"].value="Сверка ловит расхождение между сметой по позициям и фондом «бригада×срок×ставка»; это контроль порядка суммы, в ВСЕГО не входит."
    sm.row_dimensions[row].height=26
    sm.freeze_panes="A4"
    d=DataValidation(type="list",formula1=rng("A"),allow_blank=True); sm.add_data_validation(d); d.add(f"C4:C{t2}")
    # Тип ценового источника — дропдаун (оффер/объявление/прайс/приманка-от).
    dt=DataValidation(type="list",formula1='"оффер,объявление,прайс,приманка-от"',allow_blank=True); sm.add_data_validation(dt); dt.add(f"I{f1}:I{l2}")

    # Материалы
    mt=wb.create_sheet("Материалы (ведомость)"); mt.sheet_properties.tabColor="ED7D31"
    for i,w in enumerate([4,34,8,9,8,13,13,14,24,13,14,15,13,20],start=1): mt.column_dimensions[chr(64+i)].width=w
    title_rows(mt,14,f"ВЕДОМОСТЬ МАТЕРИАЛОВ (покупает собственник){head_title}{head_city}","Кол-во — по замерам. Запас % — на подрезку/бой. Цены эконом/комфорт — из живых прайсов; рядом обязательны источник/URL и дата. Суммы считаются сами.")
    header(mt,["№","Материал","Ед. изм.","Кол-во","Запас %","Кол-во с запасом",f"Цена эконом, {cur}",f"Цена комфорт, {cur}","Источник/URL","Дата",f"Сумма эконом, {cur}",f"Сумма комфорт, {cur}","Кто покупает","Примечание"])
    mt.row_dimensions[3].height=32; r=4; mf=r
    for i,(name,unit,waste) in enumerate(materials,start=1):
        SC(mt.cell(row=r,column=1,value=i),F(9),None,AL("center")); SC(mt.cell(row=r,column=2,value=name),F(9),None,AL("left"))
        SC(mt.cell(row=r,column=3,value=unit),F(9),None,AL("center")); ghost(mt,r,4,"кол-во")
        if waste: inp(mt,r,5,waste/100,numfmt='0%')
        else: ghost(mt,r,5,"%",numfmt='0%')
        SC(mt.cell(row=r,column=6,value=f"=N(D{r})*(1+N(E{r}))"),F(9),CALC,AL("center"),numfmt='#,##0.0')
        ghost(mt,r,7,cur); ghost(mt,r,8,cur)
        ghost(mt,r,9,"URL источника",numfmt=None,h="left"); ghost(mt,r,10,"дд.мм.гггг",numfmt=None)  # провенанс (Fix 2)
        SC(mt.cell(row=r,column=11,value=f"=F{r}*N(G{r})"),F(9,True),CALC,AL("center"),numfmt=MONEY)
        SC(mt.cell(row=r,column=12,value=f"=F{r}*N(H{r})"),F(9,True),CALC,AL("center"),numfmt=MONEY)
        SC(mt.cell(row=r,column=13,value="Собственник"),F(9),None,AL("center")); SC(mt.cell(row=r,column=14,value=""),F(9),None,AL("center"))
        mt.row_dimensions[r].height=18; r+=1
    ml=r-1
    mt.merge_cells(f"A{r}:J{r}"); SC(mt[f"A{r}"],F(11,True),TOTAL,AL("right")); mt[f"A{r}"].value="ИТОГО МАТЕРИАЛЫ (ориентир)"
    SC(mt.cell(row=r,column=11,value=f"=SUM(K{mf}:K{ml})"),F(11,True),TOTAL,AL("center"),numfmt=MONEY)
    SC(mt.cell(row=r,column=12,value=f"=SUM(L{mf}:L{ml})"),F(11,True),TOTAL,AL("center"),numfmt=MONEY)
    for c in (13,14): SC(mt.cell(row=r,column=c),F(9),TOTAL)
    mt.freeze_panes="A4"; mt.auto_filter.ref=f"A3:N{ml}"
    d=DataValidation(type="list",formula1=rng("A"),allow_blank=True); mt.add_data_validation(d); d.add(f"C4:C{ml}")

    # Сводка (Fix 1): паспорт книги + предупреждение о единственном носителе данных.
    sv=wb.create_sheet("Сводка"); sv.sheet_properties.tabColor=NAVY
    for i,w in enumerate([26,60],start=1): sv.column_dimensions[chr(64+i)].width=w
    sv.merge_cells("A1:B1"); SC(sv["A1"],F(13,True,WHITE),NAVY,AL("center"),bd=False)
    sv["A1"].value=f"СВОДКА ПО КНИГЕ{head_title}{head_city}"; sv.row_dimensions[1].height=32
    rows_sv=[("Объект",title or "—"),("Город",city or "—"),("Валюта",cur),
             ("Дата сборки",date.today().isoformat())]
    r=3
    for k,v in rows_sv:
        SC(sv.cell(row=r,column=1,value=k),F(10,True),SUB,AL("left"))
        SC(sv.cell(row=r,column=2,value=v),F(10),None,AL("left")); sv.row_dimensions[r].height=22; r+=1
    sv.merge_cells(f"A{r}:B{r}"); SC(sv[f"A{r}"],F(10,True,"9C0006"),"FFC7CE",AL("left"))
    sv[f"A{r}"].value="ВНИМАНИЕ: эта книга — единственный носитель данных. Повторная сборка генератором без флага --force запрещена (иначе внесённые цены/замеры будут затёрты)."
    sv.row_dimensions[r].height=46
    # B2: параметры расчёта мусора — фикс. ячейки B9/B10 (см. SV_COEF/SV_TRUCK). Предупреждение выше = строка 7.
    sv.merge_cells("A8:B8"); SC(sv["A8"],F(10,True),SUB,AL("left")); sv["A8"].value="Параметры расчёта мусора (строка «Вынос и вывоз мусора» в Смете)"
    SC(sv["A9"],F(10),None,AL("left")); sv["A9"].value="Мусор, м³ на 1 м² пола (демонтаж)"
    inp(sv,9,2,0.15,numfmt='#,##0.00')
    SC(sv["A10"],F(10),None,AL("left")); sv["A10"].value="Вместимость рейса, м³"
    inp(sv,10,2,5,numfmt='#,##0')

    order=["Сводка","Замеры — проёмы","Замеры — стены и пол","Тесты замеров","Решения и проверки","Смета","Материалы (ведомость)","Справочники"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99); wb.active=0
    wb.save(out_path)
    return {"out":out_path,"sheets":wb.sheetnames}


def load_config(args) -> dict:
    cfg={}
    if args.config:
        with open(args.config,encoding="utf-8") as fh: cfg=json.load(fh)
    if args.city: cfg["city"]=args.city
    if args.currency: cfg["currency"]=args.currency
    if args.title: cfg["title"]=args.title
    return cfg


def gsheet_bridge_md(out_path, city="", title=""):
    """Вариант 1 (ручной мост): инструкция переноса .xlsx в Google Таблицы + матрица ролей.
    Программно таблица не создаётся (нужна Google-авторизация — отдельный этап)."""
    fname=os.path.basename(out_path)
    head=(title or "Объект")+((" · "+city) if city else "")
    return f"""# Google Таблица для совместной работы — {head}

Файл сметы собран: **{fname}**. Программно Google-таблица пока НЕ создаётся
(нужна Google-авторизация — отдельный этап). Ниже ручной мост: ~5 минут, без логинов в скрипте.

## Перенос в Google Таблицы
1. Загрузи `{fname}` в Google Drive (drive.google.com → «Создать» → «Загрузка файлов»).
2. ПКМ по файлу → «Открыть с помощью» → «Google Таблицы».
3. Для нативного Google-файла (а не .xlsx-обёртки): в открытой таблице «Файл» → «Сохранить как Google Таблицы».
4. ЗАЩИТИ формулы: «Данные» → «Защищённые листы и диапазоны» → закрой от правки ячейки формул
   (Сумма/ИТОГО/ВСЕГО/нетто/периметр, лист «Тесты замеров»), лист «Справочники», параметры «Сводки».
   Разрешить правку — только себе.
5. РАЗДАЙ доступы по матрице ниже: «Настройки доступа» → добавь людей по email с нужной ролью.
6. ПРИВАТНОСТЬ: не открывай доступ «всем, у кого есть ссылка»; адрес и тех.паспорт в общий файл не пиши (обезличь — «Объект А»).

## Матрица ролей
| Роль | Доступ | Правит | Не трогает |
|---|---|---|---|
| Владелец квартиры | Редактор | Решения, бюджет | формулы, Справочники |
| Дизайнер | Редактор (Комментатор на Смете) | Решения (планировка, класс материалов) | цены, формулы |
| Прораб | Редактор | Замеры, цены, Тесты | Справочники |
| Мастер | Комментатор (или Редактор на Замерах) | предложения, замеры | смету целиком |
| Посуточный оператор | Читатель/Комментатор | — | всё (смотрит итог/сроки) |

## Проверка переноса — чек-лист A2 (заполни при первой конвертации)
- [ ] Формулы Сумма/ИТОГО/ВСЕГО считаются (не #ERROR, не текст)
- [ ] «Стены нетто = брутто − проёмы» считается
- [ ] Выпадающие списки работают (Ед. изм., Тип цены, Помещение)
- [ ] Цвет «Тестов» перенёсся ИЛИ статус читается по числовой колонке «Уровень»
- [ ] Гейт «цена без источника» краснеет при цене без URL
- [ ] Формат валюты отображается
- [ ] Защищённые диапазоны установлены — соредактор не правит формулы

Заметил деградацию переноса? Запиши в known-exceptions скила (что именно поплыло в Google Sheets).
"""


def main():
    ap=argparse.ArgumentParser(description="Сборщик Excel-сметы на ремонт квартиры (без хардкода региона и цен).")
    ap.add_argument("--out",required=True,help="Путь к выходному .xlsx")
    ap.add_argument("--city",default="",help="Город/регион объекта")
    ap.add_argument("--currency",default="₸",help="Валюта (₸, ₽ и т.д.)")
    ap.add_argument("--title",default="",help="Заголовок объекта (обезличенный, напр. «Объект А»)")
    ap.add_argument("--config",default="",help="JSON с rooms/openings/materials")
    ap.add_argument("--force",action="store_true",help="Перезаписать существующий --out (иначе сборка прервётся, чтобы не затереть данные)")
    ap.add_argument("--format",choices=["xlsx","gsheet"],default="xlsx",
                    help="xlsx — локальный файл; gsheet — тот же .xlsx + инструкция-мост в Google Таблицы и матрица ролей (Вариант 1)")
    args=ap.parse_args()
    # DATA-LOSS GUARD (Fix 1): не перезаписывать существующую книгу молча.
    if os.path.exists(args.out) and not args.force:
        print(f"ОШИБКА: файл уже существует: {args.out}\n"
              "Книга — единственный носитель данных (внесённые цены/замеры). "
              "Перезапись затрёт их. Укажи другой --out или добавь --force для осознанной перезаписи.",
              file=sys.stderr)
        sys.exit(1)
    if not args.city and not args.config:
        print("Внимание: город не указан. Укажи --city или --config; по умолчанию регион НЕ подставляется.",file=sys.stderr)
    res=build(load_config(args),args.out)
    print("Сохранено:",res["out"])
    print("Листы:",", ".join(res["sheets"]))
    print("Цены не заполнены: внеси живые расценки под город и текущую дату.")
    if args.format=="gsheet":
        guide=os.path.splitext(args.out)[0]+" — Google-Таблица (инструкция).md"
        with open(guide,"w",encoding="utf-8") as fh:
            fh.write(gsheet_bridge_md(args.out,args.city,args.title))
        print("Формат: Google Таблица (Вариант 1, ручной мост).")
        print("Инструкция + матрица ролей:",guide)
        print("Программное создание Google-таблицы — отдельный этап (нужна Google-авторизация).")


if __name__=="__main__":
    main()
