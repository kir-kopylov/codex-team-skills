#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Семантика Яндекс.Директа для посуточной аренды (STR): книга-ядро в .xlsx.

Не хардкодит город, region_id и пороги: всё приходит из config-JSON.
Частоты Вордстата скрипт НЕ выдумывает — они собираются отдельно (браузер
пользователя или вручную) и вливаются подкомандой `verdicts` из progress.jsonl.

Подкоманды:
    build       собрать книгу из config-JSON
                python3 semantika.py build --config cfg.json --out core.xlsx [--force]
    verdicts    влить частоты из progress.jsonl, проставить вердикты по порогам
                python3 semantika.py verdicts --config cfg.json --progress p.jsonl --out cfg2.json
    minus-check проверить минус-слова против собственных масок (стем-эвристика)
                python3 semantika.py minus-check --config cfg.json [--out cfg2.json]
    merge       дозалить новые маски/кластеры/минусы в существующую книгу
                python3 semantika.py merge --config new.json --into core.xlsx

config.json (обезличенный пример):
{
  "city": "Город N",
  "region_id": null,
  "thresholds": {"low": null, "high": null, "calibrated": false},
  "geo_clusters": [
    {"id": "C1", "name": "Центр", "streets": ["ул. Примерная 1"],
     "landmarks": ["главная площадь"], "geo_checked": false, "note": ""}
  ],
  "masks": [
    {"text": "квартиры посуточно город n", "group": "ядро", "cluster": "",
     "freq": null, "freq_date": null, "verdict": null, "note": ""},
    {"text": "квартира на сутки город n центр", "group": "гео", "cluster": "C1",
     "freq": null, "freq_date": null, "verdict": null, "note": ""}
  ],
  "minus_words": [
    {"word": "длительно", "status": "ок", "conflicts_with": [], "decision": ""}
  ],
  "tails": [],
  "campaign_overrides": []
}

progress.jsonl (одна строка = одна снятая частота, append-only):
  {"mask": "квартиры посуточно город n", "freq": 3500, "date": "2026-07-24",
   "tails": [{"phrase": "квартиры посуточно город n авито", "action": ""}]}

Защита от потери данных: `build` поверх существующего файла требует --force;
`merge` делает sidecar-бэкап и НЕ трогает колонки Вердикт/Решение/Комментарий.
"""
from __future__ import annotations
import argparse
import json
import shutil
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------- стили ----------
NAVY = "1F3864"; HEAD = "2E75B6"; SUB = "EAF1FB"; GREY = "F2F2F2"; WHITE = "FFFFFF"
WARN = "FFC7CE"; SOFT = "FFEB9C"; CALC = "EDF3FA"; INPUT = "FFFDF5"
AR = "Arial"
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def F(sz=10, b=False, color="000000", it=False):
    return Font(name=AR, size=sz, bold=b, color=color, italic=it)


def FILL(c):
    return PatternFill("solid", fgColor=c)


def AL(h="left", wrap=True, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def SC(c, font=None, fillc=None, align=None, bd=True, numfmt=None):
    c.font = font or F()
    if fillc:
        c.fill = FILL(fillc)
    c.alignment = align or AL()
    if bd:
        c.border = BORDER
    if numfmt:
        c.number_format = numfmt


# ---------- лингвистическое ядро (эвристика, НЕ лемматизация) ----------
# Суффиксы длинными вперёд; срез только если остаток >= 3 символов.
_SUFFIXES = ["иями", "ями", "ами", "ого", "его", "ому", "ему", "ыми", "ими",
             "ая", "яя", "ое", "ее", "ый", "ий", "ой", "ах", "ях", "ов", "ев",
             "ей", "ом", "ем", "у", "ю", "ы", "и", "а", "я", "о", "е", "ь"]
_OPERATORS = str.maketrans("", "", '+!"[]()')

VERDICT_IN = "в кампанию"
VERDICT_MAYBE = "гипотеза"
VERDICT_OUT = "вон"


def stem(word: str) -> str:
    """Стем-эвристика: срез самого длинного словоизменительного суффикса.
    Осознанно грубая: ложные совпадения безопасны (уходят в «спорное» на ревью)."""
    w = word.lower().replace("ё", "е")
    if len(w) <= 4:
        return w
    for suf in _SUFFIXES:
        if w.endswith(suf) and len(w) - len(suf) >= 3:
            return w[: -len(suf)]
    return w


def tokens(phrase: str) -> list[str]:
    return phrase.lower().replace("ё", "е").translate(_OPERATORS).split()


def norm_key(phrase: str) -> str:
    """Ключ дедупа: Вордстат склеивает словоформы, поэтому ключ — сортированные стемы."""
    return " ".join(sorted(stem(t) for t in tokens(phrase)))


def _stems_clash(s_minus: str, s_tok: str) -> bool:
    """Эвристика конфликта минус-слова с токеном маски. Настроена на перестраховку."""
    if s_minus == s_tok:
        return True
    # хвост словоизменения, который стем не добил: дом/доме, год/году
    if s_tok.startswith(s_minus) and len(s_tok) - len(s_minus) <= 2:
        return True
    if s_minus.startswith(s_tok) and len(s_minus) - len(s_tok) <= 2:
        return True
    # приставка: «суточн» прячется в «посуточн»
    if len(s_minus) >= 5 and s_tok.endswith(s_minus):
        return True
    return False


def check_minus_conflicts(cfg: dict) -> list[dict]:
    """Каждое не снятое минус-слово против каждого токена каждой маски.
    Конфликт => статус «спорное» + conflicts_with. Ничего не удаляет."""
    conflicts = []
    for mw in cfg.get("minus_words", []):
        if mw.get("status") == "снято":
            continue
        s_minus = stem(mw["word"])
        hit_masks = []
        for m in cfg.get("masks", []):
            for t in tokens(m["text"]):
                if _stems_clash(s_minus, stem(t)):
                    hit_masks.append(m["text"])
                    break
        if hit_masks:
            mw["status"] = "спорное"
            mw["conflicts_with"] = sorted(set(mw.get("conflicts_with", []) + hit_masks))
            conflicts.append({"word": mw["word"], "masks": hit_masks})
    return conflicts


# ---------- вердикты ----------
def load_progress(path: str) -> dict:
    """progress.jsonl -> {norm_key: {"freq":…, "date":…, "tails":[…]}}. Последняя запись побеждает."""
    out = {}
    if not path or not os.path.exists(path):
        return out
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            out[norm_key(rec["mask"])] = rec
    return out


def apply_verdicts(cfg: dict, progress: dict) -> dict:
    th = cfg.get("thresholds") or {}
    low, high = th.get("low"), th.get("high")
    for m in cfg.get("masks", []):
        rec = progress.get(norm_key(m["text"]))
        if rec:
            m["freq"] = rec.get("freq")
            m["freq_date"] = rec.get("date")
            for tail in rec.get("tails", []) or []:
                if tail.get("phrase"):
                    cfg.setdefault("tails", []).append(
                        {"phrase": tail["phrase"], "mask": m["text"],
                         "action": tail.get("action", "")})
        if m.get("freq") is None or low is None or high is None:
            continue
        f = m["freq"]
        m["verdict"] = VERDICT_OUT if f < low else (VERDICT_MAYBE if f <= high else VERDICT_IN)
    # дедуп хвостов
    seen, uniq = set(), []
    for t in cfg.get("tails", []):
        k = norm_key(t["phrase"])
        if k not in seen:
            seen.add(k)
            uniq.append(t)
    cfg["tails"] = uniq
    return cfg


# ---------- кампании ----------
# Токены, которые не различают группы (ядро STR + служебные) — не годятся в кросс-минусы.
_CORE_STEMS = {stem(w) for w in [
    "квартира", "квартиры", "снять", "аренда", "посуточно", "сутки", "жилье",
    "апартаменты", "на", "в", "с", "для", "у", "и", "по", "рядом", "недорого",
    "забронировать", "сниму", "суточно", "краткосрочная", "день"]}


def _group_ax(m: dict) -> tuple[str, str]:
    """(кампания, группа объявлений) для маски. 1 группа = 1 интент."""
    g = (m.get("group") or "прочее").strip().lower()
    if g.startswith("гео"):
        return "Гео", (m.get("cluster") or "без кластера")
    return "Ядро и спрос", m.get("group") or "прочее"


def build_campaigns(cfg: dict) -> list[dict]:
    """Раскладка масок с вердиктом «в кампанию» + детерминированная кросс-минусовка."""
    city = cfg.get("city", "")
    city_stems = {stem(t) for t in tokens(city)}
    groups: dict[tuple[str, str], list[dict]] = {}
    for m in cfg.get("masks", []):
        if m.get("verdict") != VERDICT_IN:
            continue
        groups.setdefault(_group_ax(m), []).append(m)

    stems_by_group = {}
    for key, masks in groups.items():
        st = set()
        for m in masks:
            st |= {stem(t) for t in tokens(m["text"])}
        stems_by_group[key] = st - _CORE_STEMS - city_stems

    rows = []
    for (camp, adg), masks in sorted(groups.items()):
        own = stems_by_group[(camp, adg)]
        cross = set()
        for other, st in stems_by_group.items():
            if other[0] == camp and other != (camp, adg):
                cross |= st - own
        cross_txt = ", ".join(sorted(cross)[:10])
        for m in masks:
            rows.append({"campaign": f"Поиск — {city} — {camp}",
                         "adgroup": adg, "intent": adg,
                         "mask": m["text"], "cross_minus": cross_txt,
                         "note": m.get("note", "")})
    for o in cfg.get("campaign_overrides", []) or []:
        rows.append(o)
    return rows


# ---------- сборка книги ----------
SHEETS = ["Сводка", "Маски", "Минус-слова", "Гео-кластеры", "Хвосты", "Кампании"]


def _sheet_campaigns(wb, campaigns):
    """(Пере)создаёт лист «Кампании». Лист производный — генерируется из масок,
    ручных правок не содержит, поэтому пересоздание безопасно."""
    if "Кампании" in wb.sheetnames:
        wb.remove(wb["Кампании"])
    cp = wb.create_sheet("Кампании")
    cp.sheet_properties.tabColor = NAVY
    _header(cp, ["Кампания", "Группа объявлений", "Интент группы", "Маска",
                 "Кросс-минусы группы", "Комментарий"])
    for col, w in zip("ABCDEF", [30, 22, 22, 46, 44, 24]):
        cp.column_dimensions[col].width = w
    r = 2
    if campaigns:
        for row_c in campaigns:
            vals = [row_c.get("campaign", ""), row_c.get("adgroup", ""),
                    row_c.get("intent", ""), row_c.get("mask", ""),
                    row_c.get("cross_minus", ""), row_c.get("note", "")]
            for c, v in enumerate(vals, 1):
                SC(cp.cell(row=r, column=c, value=v), F(10), None,
                   AL("left") if c in (4, 5, 6) else AL("center"))
            r += 1
    else:
        cp.merge_cells("A2:F2")
        SC(cp["A2"], F(10, True, "9C6500"), SOFT, AL("left"))
        cp["A2"].value = ("Лист заполнится после вердиктов: сюда попадают только маски "
                          "с вердиктом «в кампанию» (semantika.py verdicts → build/merge).")
        cp.row_dimensions[2].height = 30
    cp.freeze_panes = "A2"
    return cp


def _header(ws, headers, row=1):
    for i, h in enumerate(headers, start=1):
        SC(ws.cell(row=row, column=i, value=h), F(10, True, WHITE), HEAD, AL("center"))
    ws.row_dimensions[row].height = 26


def _dv(ws, values, cellrange):
    d = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(d)
    d.add(cellrange)


def build(cfg: dict, out_path: str) -> dict:
    city = cfg.get("city", "").strip()
    th = cfg.get("thresholds") or {}
    low, high, calibrated = th.get("low"), th.get("high"), th.get("calibrated")
    masks = cfg.get("masks", [])
    minus = cfg.get("minus_words", [])
    clusters = cfg.get("geo_clusters", [])
    tails = cfg.get("tails", [])
    campaigns = build_campaigns(cfg)

    wb = Workbook()

    # Сводка — паспорт прогона (нужна merge-режиму и человеку)
    sv = wb.active
    sv.title = "Сводка"
    sv.sheet_properties.tabColor = NAVY
    sv.column_dimensions["A"].width = 34
    sv.column_dimensions["B"].width = 70
    sv.merge_cells("A1:B1")
    SC(sv["A1"], F(13, True, WHITE), NAVY, AL("center"), bd=False)
    sv["A1"].value = f"СЕМАНТИКА ДИРЕКТА — ПОСУТОЧНАЯ АРЕНДА · {city or '—'}"
    sv.row_dimensions[1].height = 32
    freq_dates = sorted({m["freq_date"] for m in masks if m.get("freq_date")})
    n_freq = sum(1 for m in masks if m.get("freq") is not None)
    thr_txt = (f"низкий {low} / высокий {high}" if low is not None and high is not None
               else "не заданы")
    if not calibrated:
        thr_txt += "  [requires recheck] — откалибровать на первом прогоне частот"
    rows_sv = [
        ("Город", city or "—"),
        ("region_id Вордстата", cfg.get("region_id") if cfg.get("region_id") is not None
         else "—  [requires recheck] — зафиксировать при первом прогоне"),
        ("Пороги вердиктов (показов/мес)", thr_txt),
        ("Масок всего / с частотой", f"{len(masks)} / {n_freq}"),
        ("Даты сбора частот", ", ".join(freq_dates) or "частоты не собирались"),
        ("Гео-кластеры", ", ".join(f"{c['id']} {c.get('name', '')}".strip() for c in clusters) or "—"),
    ]
    r = 3
    for k, v in rows_sv:
        SC(sv.cell(row=r, column=1, value=k), F(10, True), SUB, AL("left"))
        SC(sv.cell(row=r, column=2, value=v), F(10), None, AL("left"))
        sv.row_dimensions[r].height = 22
        r += 1
    r += 1
    sv.merge_cells(f"A{r}:B{r}")
    SC(sv[f"A{r}"], F(10, True, "9C0006"), WARN, AL("left"))
    sv[f"A{r}"].value = ("ВНИМАНИЕ: после ручных правок (вердикты, решения по минусам) книга — "
                         "единственный носитель этих правок. Пересборка build поверх неё требует --force "
                         "и затирает ручное; дозаливка — только через merge (он делает бэкап и не трогает ручные колонки).")
    sv.row_dimensions[r].height = 58

    # Маски
    ms = wb.create_sheet("Маски")
    ms.sheet_properties.tabColor = "548235"
    _header(ms, ["№", "Маска", "Группа спроса", "Гео-кластер", "Частота",
                 "Дата частоты", "Вердикт", "Комментарий (1 строка)", "Норм. ключ"])
    for col, w in zip("ABCDEFGHI", [5, 50, 22, 13, 11, 13, 14, 44, 40]):
        ms.column_dimensions[col].width = w
    r = 2
    for i, m in enumerate(masks, start=1):
        vals = [i, m["text"], m.get("group", ""), m.get("cluster", ""),
                m.get("freq"), m.get("freq_date", ""), m.get("verdict", ""),
                m.get("note", ""), norm_key(m["text"])]
        for c, v in enumerate(vals, 1):
            cell = ms.cell(row=r, column=c, value=v)
            SC(cell, F(10), INPUT if c in (5, 7) else None,
               AL("left") if c in (2, 8, 9) else AL("center"),
               numfmt="#,##0" if c == 5 else None)
        ms.row_dimensions[r].height = 18
        r += 1
    last = max(r - 1, 2)
    ms.freeze_panes = "C2"
    ms.auto_filter.ref = f"A1:I{last}"
    _dv(ms, [VERDICT_IN, VERDICT_MAYBE, VERDICT_OUT], f"G2:G{last}")
    ms.conditional_formatting.add(
        f"A2:I{last}",
        FormulaRule(formula=['AND($G2<>"",$E2="")'], fill=FILL(WARN)))

    # Минус-слова
    mn = wb.create_sheet("Минус-слова")
    mn.sheet_properties.tabColor = "C55A11"
    _header(mn, ["№", "Минус-слово", "Статус", "Конфликтует с маской", "Решение", "Комментарий"])
    for col, w in zip("ABCDEF", [5, 20, 12, 46, 46, 30]):
        mn.column_dimensions[col].width = w
    r = 2
    for i, mw in enumerate(minus, start=1):
        vals = [i, mw["word"], mw.get("status", "ок"),
                "; ".join(mw.get("conflicts_with", [])), mw.get("decision", ""),
                mw.get("note", "")]
        for c, v in enumerate(vals, 1):
            SC(mn.cell(row=r, column=c, value=v), F(10),
               INPUT if c == 5 else None,
               AL("left") if c in (4, 5, 6) else AL("center"))
        mn.row_dimensions[r].height = 18
        r += 1
    last = max(r - 1, 2)
    mn.freeze_panes = "A2"
    mn.auto_filter.ref = f"A1:F{last}"
    _dv(mn, ["ок", "спорное", "снято"], f"C2:C{last}")
    mn.conditional_formatting.add(
        f"A2:F{last}",
        FormulaRule(formula=['AND($C2="спорное",$E2="")'], fill=FILL(SOFT)))

    # Гео-кластеры
    gk = wb.create_sheet("Гео-кластеры")
    gk.sheet_properties.tabColor = HEAD
    _header(gk, ["ID", "Кластер/район", "Улицы (адреса)", "Ориентиры",
                 "Гео-проверка", "Дата заливки", "Комментарий"])
    for col, w in zip("ABCDEFG", [6, 22, 40, 44, 13, 13, 26]):
        gk.column_dimensions[col].width = w
    r = 2
    for c_ in clusters:
        vals = [c_["id"], c_.get("name", ""), "; ".join(c_.get("streets", [])),
                "; ".join(c_.get("landmarks", [])),
                "да" if c_.get("geo_checked") else "нет",
                c_.get("added_date", ""), c_.get("note", "")]
        for c, v in enumerate(vals, 1):
            SC(gk.cell(row=r, column=c, value=v), F(10), None,
               AL("left") if c in (3, 4, 7) else AL("center"))
        gk.row_dimensions[r].height = 20
        r += 1
    gk.freeze_panes = "A2"
    _dv(gk, ["да", "нет"], f"E2:E{max(r - 1, 2)}")

    # Хвосты
    tl = wb.create_sheet("Хвосты")
    tl.sheet_properties.tabColor = "A6300F"
    _header(tl, ["Фраза-хвост из Вордстата", "У какой маски увидена", "Действие", "Комментарий"])
    for col, w in zip("ABCD", [50, 44, 16, 26]):
        tl.column_dimensions[col].width = w
    r = 2
    for t in tails:
        vals = [t["phrase"], t.get("mask", ""), t.get("action", ""), t.get("note", "")]
        for c, v in enumerate(vals, 1):
            SC(tl.cell(row=r, column=c, value=v), F(10), None,
               AL("left") if c in (1, 2, 4) else AL("center"))
        r += 1
    last = max(r - 1, 2)
    tl.freeze_panes = "A2"
    _dv(tl, ["в минус-лист", "игнор", "в маски"], f"C2:C{last}")

    # Кампании
    _sheet_campaigns(wb, campaigns)

    wb.save(out_path)
    return {"out": out_path, "sheets": wb.sheetnames,
            "masks": len(masks), "campaign_rows": len(campaigns)}


# ---------- merge (дозаливка) ----------
def merge(cfg: dict, xlsx_path: str) -> dict:
    """Добавляет в существующую книгу ТОЛЬКО новые строки (дедуп по норм-ключу).
    Существующие строки, включая ручные Вердикт/Решение/Комментарий, не трогает."""
    backup = (os.path.splitext(xlsx_path)[0]
              + ".backup-" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx")
    shutil.copy2(xlsx_path, backup)
    wb = load_workbook(xlsx_path)
    for s in SHEETS:
        if s not in wb.sheetnames:
            raise SystemExit(f"ОШИБКА: в книге нет листа «{s}» — это не книга semantika.py")

    ms = wb["Маски"]
    existing = set()
    for row in ms.iter_rows(min_row=2):
        text = row[1].value
        key = row[8].value or (norm_key(text) if text else None)
        if key:
            existing.add(key)
    added_masks = 0
    r = ms.max_row + 1
    for m in cfg.get("masks", []):
        key = norm_key(m["text"])
        if key in existing:
            continue
        existing.add(key)
        vals = [r - 1, m["text"], m.get("group", ""), m.get("cluster", ""),
                m.get("freq"), m.get("freq_date", ""), m.get("verdict", ""),
                m.get("note", ""), key]
        for c, v in enumerate(vals, 1):
            SC(ms.cell(row=r, column=c, value=v), F(10),
               INPUT if c in (5, 7) else None,
               AL("left") if c in (2, 8, 9) else AL("center"),
               numfmt="#,##0" if c == 5 else None)
        r += 1
        added_masks += 1

    mn = wb["Минус-слова"]
    existing_minus = {stem(row[1].value) for row in mn.iter_rows(min_row=2) if row[1].value}
    added_minus = 0
    r = mn.max_row + 1
    for mw in cfg.get("minus_words", []):
        if stem(mw["word"]) in existing_minus:
            continue
        existing_minus.add(stem(mw["word"]))
        vals = [r - 1, mw["word"], mw.get("status", "ок"),
                "; ".join(mw.get("conflicts_with", [])), mw.get("decision", ""), ""]
        for c, v in enumerate(vals, 1):
            SC(mn.cell(row=r, column=c, value=v), F(10), INPUT if c == 5 else None,
               AL("left") if c in (4, 5, 6) else AL("center"))
        r += 1
        added_minus += 1

    gk = wb["Гео-кластеры"]
    existing_cl = {row[0].value for row in gk.iter_rows(min_row=2) if row[0].value}
    added_cl = 0
    r = gk.max_row + 1
    today = datetime.now().strftime("%Y-%m-%d")
    for c_ in cfg.get("geo_clusters", []):
        if c_["id"] in existing_cl:
            continue
        vals = [c_["id"], c_.get("name", ""), "; ".join(c_.get("streets", [])),
                "; ".join(c_.get("landmarks", [])),
                "да" if c_.get("geo_checked") else "нет", today, c_.get("note", "")]
        for c, v in enumerate(vals, 1):
            SC(gk.cell(row=r, column=c, value=v), F(10), None,
               AL("left") if c in (3, 4, 7) else AL("center"))
        r += 1
        added_cl += 1

    # «Кампании» — производный лист: пересобираем из ИТОГОВОГО состояния «Масок»
    # (включая ручные вердикты книги и только что дозалитые маски), иначе новый
    # кластер с вердиктом «в кампанию» молча не попадёт в кампании.
    city = ""
    for row in wb["Сводка"].iter_rows(min_col=1, max_col=2):
        if row[0].value == "Город" and row[1].value and row[1].value != "—":
            city = str(row[1].value)
    masks_state = []
    for row in ms.iter_rows(min_row=2):
        if not row[1].value:
            continue
        masks_state.append({"text": row[1].value, "group": row[2].value or "",
                            "cluster": row[3].value or "", "verdict": row[6].value,
                            "note": row[7].value or ""})
    campaigns = build_campaigns({"city": city, "masks": masks_state,
                                 "campaign_overrides": cfg.get("campaign_overrides", [])})
    _sheet_campaigns(wb, campaigns)

    wb.save(xlsx_path)
    return {"backup": backup, "added_masks": added_masks,
            "added_minus": added_minus, "added_clusters": added_cl,
            "campaign_rows": len(campaigns)}


# ---------- CLI ----------
def _load_cfg(path: str) -> dict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _save_cfg(cfg: dict, path: str):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, ensure_ascii=False, indent=1)


def main():
    ap = argparse.ArgumentParser(description="Семантика Директа для STR: build / verdicts / minus-check / merge.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="собрать книгу из config-JSON")
    b.add_argument("--config", required=True)
    b.add_argument("--out", required=True)
    b.add_argument("--force", action="store_true",
                   help="перезаписать существующую книгу (затирает ручные правки!)")

    v = sub.add_parser("verdicts", help="влить частоты и проставить вердикты")
    v.add_argument("--config", required=True)
    v.add_argument("--progress", default="", help="progress.jsonl со снятыми частотами")
    v.add_argument("--out", required=True, help="куда писать обновлённый config (не in-place)")

    mc = sub.add_parser("minus-check", help="минус-слова против собственных масок (эвристика)")
    mc.add_argument("--config", required=True)
    mc.add_argument("--out", default="", help="куда писать config со статусами «спорное»")

    mg = sub.add_parser("merge", help="дозалить новые маски/кластеры/минусы в книгу")
    mg.add_argument("--config", required=True)
    mg.add_argument("--into", required=True, help="существующая книга .xlsx")

    args = ap.parse_args()

    if args.cmd == "build":
        if os.path.exists(args.out) and not args.force:
            print(f"ОШИБКА: файл уже существует: {args.out}\n"
                  "Книга — носитель ручных правок (вердикты, решения по минусам). "
                  "Перезапись затрёт их: для дозаливки используй merge, "
                  "для осознанной пересборки добавь --force.", file=sys.stderr)
            sys.exit(1)
        res = build(_load_cfg(args.config), args.out)
        print("Сохранено:", res["out"])
        print("Листы:", ", ".join(res["sheets"]))
        print(f"Масок: {res['masks']}; строк в «Кампании»: {res['campaign_rows']}")
        if not res["campaign_rows"]:
            print("«Кампании» пусты: вердиктов «в кампанию» ещё нет (сначала частоты → verdicts).")

    elif args.cmd == "verdicts":
        cfg = _load_cfg(args.config)
        th = cfg.get("thresholds") or {}
        if th.get("low") is None or th.get("high") is None:
            print("Внимание: пороги не заданы — частоты вольются, вердикты НЕ проставятся. "
                  "Задай thresholds.low/high в config (калибровка на первом прогоне).",
                  file=sys.stderr)
        cfg = apply_verdicts(cfg, load_progress(args.progress))
        _save_cfg(cfg, args.out)
        n = sum(1 for m in cfg["masks"] if m.get("verdict"))
        nf = sum(1 for m in cfg["masks"] if m.get("freq") is not None)
        print(f"Частот влито: {nf}; вердиктов проставлено: {n}; хвостов: {len(cfg.get('tails', []))}")
        print("Обновлённый config:", args.out)

    elif args.cmd == "minus-check":
        cfg = _load_cfg(args.config)
        conflicts = check_minus_conflicts(cfg)
        if conflicts:
            print(f"Найдено конфликтов (стем-эвристика, возможны ложные — решает человек): {len(conflicts)}")
            for c in conflicts:
                print(f"  «{c['word']}»  ↔  " + "; ".join(c["masks"][:4])
                      + (" …" if len(c["masks"]) > 4 else ""))
            print("Эти минус-слова помечены «спорное» — заполни «Решение» по каждому.")
        else:
            print("Конфликтов минус-листа с масками не найдено (эвристика).")
        if args.out:
            _save_cfg(cfg, args.out)
            print("Config со статусами:", args.out)

    elif args.cmd == "merge":
        res = merge(_load_cfg(args.config), args.into)
        print("Бэкап:", res["backup"])
        print(f"Добавлено: масок {res['added_masks']}, минус-слов {res['added_minus']}, "
              f"кластеров {res['added_clusters']}. Существующие строки не тронуты.")


if __name__ == "__main__":
    main()
