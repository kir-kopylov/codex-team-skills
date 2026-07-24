# -*- coding: utf-8 -*-
"""Скил-тест str-direct-semantika: инварианты книги, стем-эвристики, вердиктов,
кампаний и дозаливки. Регресс-фикстура — из реального прогона (минусы
ночь/недорого/суточно/дом/год убивали собственные маски)."""
import importlib.util
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT = Path(__file__).resolve().parents[1] / "plugins/team-skills/skills/str-direct-semantika/scripts/semantika.py"
SHEETS = ["Сводка", "Маски", "Минус-слова", "Гео-кластеры", "Хвосты", "Кампании"]


def _mod():
    spec = importlib.util.spec_from_file_location("semantika", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _cfg():
    return {
        "city": "Город N",
        "region_id": None,
        "thresholds": {"low": 10, "high": 100, "calibrated": False},
        "geo_clusters": [
            {"id": "C1", "name": "Центр", "streets": ["ул. Примерная 1"],
             "landmarks": ["главная площадь"], "geo_checked": True, "note": ""},
            {"id": "C2", "name": "Вокзал", "streets": ["ул. Тестовая 2"],
             "landmarks": ["жд вокзал"], "geo_checked": True, "note": ""},
        ],
        "masks": [
            {"text": "квартиры посуточно город n", "group": "ядро", "cluster": "",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "квартира на ночь недорого город n", "group": "сроки", "cluster": "",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "квартира посуточно в новом доме город n", "group": "типы", "cluster": "",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "снять квартиру на новый год город n", "group": "события", "cluster": "",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "квартиры суточно город n", "group": "ядро", "cluster": "",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "квартира посуточно город n центр", "group": "гео", "cluster": "C1",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
            {"text": "квартира на сутки город n жд вокзал", "group": "гео", "cluster": "C2",
             "freq": None, "freq_date": None, "verdict": None, "note": ""},
        ],
        "minus_words": [
            {"word": "ночь", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "недорого", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "суточно", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "дом", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "год", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "длительно", "status": "ок", "conflicts_with": [], "decision": ""},
            {"word": "помесячно", "status": "ок", "conflicts_with": [], "decision": ""},
        ],
        "tails": [],
        "campaign_overrides": [],
    }


# ---------- лингвистическое ядро ----------

def test_dedup_wordforms():
    m = _mod()
    assert m.norm_key("квартиры посуточно город n") == m.norm_key("квартира посуточно город n")
    assert m.norm_key("снять квартиру") != m.norm_key("снять комнату")


def test_minus_selfconflicts_regression():
    """Реальный кейс: минусы ночь/недорого/суточно/дом/год режут собственные маски."""
    m = _mod()
    cfg = _cfg()
    conflicts = m.check_minus_conflicts(cfg)
    flagged = {c["word"] for c in conflicts}
    for w in ["ночь", "недорого", "суточно", "дом", "год"]:
        assert w in flagged, f"минус «{w}» должен конфликтовать с собственной маской"
    by_word = {mw["word"]: mw for mw in cfg["minus_words"]}
    for w in ["ночь", "недорого", "суточно", "дом", "год"]:
        assert by_word[w]["status"] == "спорное"
        assert by_word[w]["conflicts_with"]
    assert by_word["длительно"]["status"] == "ок", "чистый минус не должен помечаться"
    assert by_word["помесячно"]["status"] == "ок"


def test_minus_check_never_deletes():
    m = _mod()
    cfg = _cfg()
    n = len(cfg["minus_words"])
    m.check_minus_conflicts(cfg)
    assert len(cfg["minus_words"]) == n, "minus-check не имеет права удалять минус-слова"


# ---------- вердикты ----------

def test_verdicts_two_threshold():
    m = _mod()
    cfg = _cfg()
    progress = {
        m.norm_key("квартиры посуточно город n"): {"freq": 500, "date": "2026-01-01"},
        m.norm_key("квартира посуточно город n центр"): {"freq": 50, "date": "2026-01-01"},
        m.norm_key("квартира на сутки город n жд вокзал"): {"freq": 5, "date": "2026-01-01"},
    }
    cfg = m.apply_verdicts(cfg, progress)
    by_text = {x["text"]: x for x in cfg["masks"]}
    assert by_text["квартиры посуточно город n"]["verdict"] == m.VERDICT_IN
    assert by_text["квартира посуточно город n центр"]["verdict"] == m.VERDICT_MAYBE
    assert by_text["квартира на сутки город n жд вокзал"]["verdict"] == m.VERDICT_OUT
    assert by_text["квартиры суточно город n"]["verdict"] is None, "без частоты вердикта нет"


def test_verdicts_collect_tails():
    m = _mod()
    cfg = _cfg()
    progress = {m.norm_key("квартиры посуточно город n"): {
        "freq": 500, "date": "2026-01-01",
        "tails": [{"phrase": "квартиры посуточно город n длительно", "action": ""}]}}
    cfg = m.apply_verdicts(cfg, progress)
    assert any("длительно" in t["phrase"] for t in cfg["tails"])


# ---------- кампании ----------

def test_campaigns_only_verdict_in_and_cross_minus():
    m = _mod()
    cfg = _cfg()
    for x in cfg["masks"]:
        if x["group"] == "гео":
            x["verdict"] = m.VERDICT_IN
    rows = m.build_campaigns(cfg)
    assert rows, "гео-маски с вердиктом «в кампанию» должны попасть в кампании"
    assert all(r["campaign"].startswith("Поиск — Город N") for r in rows)
    adgroups = {r["adgroup"] for r in rows}
    assert adgroups == {"C1", "C2"}, "1 группа объявлений = 1 гео-кластер"
    c1 = next(r for r in rows if r["adgroup"] == "C1")
    assert "вокзал" in c1["cross_minus"], "дифференциатор сиблинга должен уйти в кросс-минусы"
    without_verdict = [r for r in rows if r["mask"] == "квартиры посуточно город n"]
    assert not without_verdict, "маска без вердикта «в кампанию» не должна попасть в лист"


# ---------- книга ----------

def test_build_sheets_and_manual_flag(tmp_path):
    m = _mod()
    out = tmp_path / "core.xlsx"
    m.build(_cfg(), str(out))
    wb = load_workbook(out)
    for s in SHEETS:
        assert s in wb.sheetnames, f"нет листа {s}"
    sv_text = " ".join(str(c.value) for row in wb["Сводка"].iter_rows() for c in row if c.value)
    assert "[requires recheck]" in sv_text, "неоткалиброванный порог обязан быть помечен"
    ms = wb["Маски"]
    assert ms.cell(row=1, column=9).value == "Норм. ключ"
    assert ms.cell(row=2, column=9).value == m.norm_key(ms.cell(row=2, column=2).value)


def test_data_loss_guard(tmp_path):
    cfg_p = tmp_path / "cfg.json"
    cfg_p.write_text(json.dumps(_cfg(), ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "guard.xlsx"
    r1 = subprocess.run([sys.executable, str(SCRIPT), "build", "--config", str(cfg_p),
                         "--out", str(out)], capture_output=True)
    assert r1.returncode == 0 and out.exists(), r1.stderr
    r2 = subprocess.run([sys.executable, str(SCRIPT), "build", "--config", str(cfg_p),
                         "--out", str(out)], capture_output=True)
    assert r2.returncode != 0, "перезапись без --force должна прерываться"
    r3 = subprocess.run([sys.executable, str(SCRIPT), "build", "--config", str(cfg_p),
                         "--out", str(out), "--force"], capture_output=True)
    assert r3.returncode == 0, r3.stderr


def test_merge_preserves_manual_and_backups(tmp_path):
    m = _mod()
    out = tmp_path / "core.xlsx"
    m.build(_cfg(), str(out))
    wb = load_workbook(out)
    ms = wb["Маски"]
    ms.cell(row=2, column=7, value="в кампанию")   # ручной вердикт
    wb["Минус-слова"].cell(row=2, column=5, value="не минусовать")  # ручное решение
    wb.save(out)
    rows_before = load_workbook(out)["Маски"].max_row

    new_cfg = {"city": "Город N",
               "geo_clusters": [{"id": "C3", "name": "Новый район", "streets": ["ул. Новая 3"],
                                 "landmarks": [], "geo_checked": True}],
               "masks": [
                   {"text": "квартира посуточно город n", "group": "ядро", "cluster": ""},
                   {"text": "квартира посуточно город n новый район", "group": "гео",
                    "cluster": "C3", "freq": 300, "verdict": "в кампанию"},
               ],
               "minus_words": [{"word": "длительно", "status": "ок"},
                               {"word": "офис", "status": "ок"}]}
    res = m.merge(new_cfg, str(out))
    assert Path(res["backup"]).exists(), "merge обязан оставить бэкап"
    assert res["added_masks"] == 1, "дубль по словоформе не должен добавляться повторно"
    assert res["added_minus"] == 1, "существующий минус не дублируется"
    assert res["added_clusters"] == 1
    wb2 = load_workbook(out)
    assert wb2["Маски"].max_row == rows_before + 1
    assert wb2["Маски"].cell(row=2, column=7).value == "в кампанию", "ручной вердикт затёрт"
    assert wb2["Минус-слова"].cell(row=2, column=5).value == "не минусовать", "ручное решение затёрто"
    # «Кампании» пересобраны из итогового состояния масок: и дозалитая маска
    # с вердиктом «в кампанию», и маска с ручным вердиктом из книги.
    camp_masks = [wb2["Кампании"].cell(row=r, column=4).value
                  for r in range(2, wb2["Кампании"].max_row + 1)]
    assert "квартира посуточно город n новый район" in camp_masks, \
        "дозалитый кластер с вердиктом «в кампанию» пропал из «Кампаний»"
    assert "квартиры посуточно город n" in camp_masks, \
        "ручной вердикт «в кампанию» из книги не попал в «Кампании»"


def test_merge_refuses_foreign_book(tmp_path):
    m = _mod()
    from openpyxl import Workbook
    alien = tmp_path / "alien.xlsx"
    Workbook().save(alien)
    try:
        m.merge({"masks": []}, str(alien))
        assert False, "merge обязан отказаться от чужой книги"
    except SystemExit:
        pass
