# -*- coding: utf-8 -*-
"""Скил-тест remont-smeta-builder: строит книгу и проверяет инварианты статически
(без LibreOffice/recalc) + data-loss-гейт через subprocess. Ловит регресс при правках
генератора. Закрывает аудит-находки: привязка замеры→смета, гейт цен, портируемый
статус, итог-0, защита от перезаписи, сверка труда (B1), мусор формулой (B2)."""
import importlib.util
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
SCRIPT = REPO / "plugins/team-skills/skills/remont-smeta-builder/scripts/build_smeta_xlsx.py"
SHEETS = ["Сводка", "Замеры — проёмы", "Замеры — стены и пол", "Тесты замеров",
          "Решения и проверки", "Смета", "Материалы (ведомость)", "Справочники"]
ERR = re.compile(r"#(REF|DIV/0|VALUE|NAME\?|N/A|NUM|NULL)")


def _build(tmp_path, **cfg):
    spec = importlib.util.spec_from_file_location("build_smeta_xlsx", SCRIPT)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    out = tmp_path / "smeta.xlsx"
    mod.build({"city": "Тест", "currency": "₸", **cfg}, str(out))
    return load_workbook(out)


def _strings(ws, upto_row=None):
    out = []
    for row in ws.iter_rows(max_row=upto_row):
        for c in row:
            if isinstance(c.value, str):
                out.append(c.value)
    return out


def test_all_sheets_present(tmp_path):
    wb = _build(tmp_path)
    for s in SHEETS:
        assert s in wb.sheetnames, f"нет листа {s}"


def test_no_error_tokens_in_formulas(tmp_path):
    wb = _build(tmp_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    assert not ERR.search(c.value), f"{ws.title}!{c.coordinate}: {c.value}"


def test_price_provenance_columns(tmp_path):
    sm = _build(tmp_path)["Смета"]
    head = _strings(sm, upto_row=3)
    assert any("Ценовой источник" in x for x in head), "нет колонки ценового источника"
    assert any("Дата котировки" in x for x in head), "нет колонки даты котировки"


def test_qty_linked_to_zameri(tmp_path):
    sm = _build(tmp_path)["Смета"]
    linked = [c.value for row in sm.iter_rows() for c in row
              if isinstance(c.value, str) and c.value.startswith("=") and "Замеры — стены и пол" in c.value]
    assert linked, "ни одно Кол-во не связано формулой с листом замеров"


def test_totals_guarded_against_empty(tmp_path):
    sm = _build(tmp_path)["Смета"]
    guarded = [c.value for row in sm.iter_rows() for c in row
               if isinstance(c.value, str) and "IF(COUNT(" in c.value]
    assert guarded, "ИТОГО/ВСЕГО не защищены IF(COUNT(...)) — будет тихий 0"


def test_price_gate_and_portable_level(tmp_path):
    ts = _build(tmp_path)["Тесты замеров"]
    texts = _strings(ts)
    assert any("без ценового источника" in t for t in texts), "нет гейта «цена без источника»"
    assert any(t == "Уровень" for t in texts), "нет числовой колонки «Уровень»"
    a3 = ts["A3"].value
    assert isinstance(a3, str) and "COUNTIF" not in a3.upper(), "баннер на wildcard-COUNTIF (непортируемо в Numbers)"


def test_completeness_excludes_add_rows(tmp_path):
    # тест полноты должен ссылаться на data-range проёмов (D4:E9 для 6 проёмов),
    # а не включать хвостовые строки «(добавить)».
    ts = _build(tmp_path)["Тесты замеров"]
    comp = [c.value for row in ts.iter_rows() for c in row
            if isinstance(c.value, str) and "Замеры — проёмы" in c.value and "D4:E" in c.value]
    assert comp, "формула полноты не использует data-range проёмов"
    assert all("E11" not in f and "E10" not in f for f in comp), "полнота цепляет строки «(добавить)»"


def test_per_row_opening_height_helpers(tmp_path):
    pr = _build(tmp_path)["Замеры — проёмы"]
    head = _strings(pr, upto_row=3)
    assert any("Высота комнаты" in x for x in head), "нет служебной колонки «Высота комнаты»"
    assert any(x == "Превышение" for x in head), "нет построчного флага «Превышение»"


def test_labor_crosscheck_block_B1(tmp_path):
    sm = _build(tmp_path)["Смета"]
    texts = _strings(sm)
    assert any("СВЕРКА ТРУДА" in t for t in texts), "нет блока сверки труда (B1)"
    assert any("Фонд по бригаде" in t for t in texts), "нет строки фонда бригады"
    assert any("Статус сверки труда" in t for t in texts), "нет статуса сверки"


def test_debris_is_formula_B2(tmp_path):
    sm = _build(tmp_path)["Смета"]
    target = None
    for row in sm.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Вынос и вывоз мусора" in c.value:
                target = c.row
                break
        if target:
            break
    assert target, "нет строки «Вынос и вывоз мусора»"
    d = sm.cell(row=target, column=4).value  # Кол-во
    assert isinstance(d, str) and d.startswith("="), f"Кол-во мусора не формула: {d!r}"
    assert "ROUNDUP" in d and "Сводка" in d, "мусор не считается из объёма/вместимости рейса"


def test_data_loss_guard(tmp_path):
    out = tmp_path / "guard.xlsx"
    r1 = subprocess.run([sys.executable, str(SCRIPT), "--out", str(out), "--city", "Тест"],
                        capture_output=True)
    assert r1.returncode == 0 and out.exists(), "первая сборка должна пройти"
    r2 = subprocess.run([sys.executable, str(SCRIPT), "--out", str(out), "--city", "Тест"],
                        capture_output=True)
    assert r2.returncode != 0, "перезапись существующего файла без --force должна прерываться"
    r3 = subprocess.run([sys.executable, str(SCRIPT), "--out", str(out), "--city", "Тест", "--force"],
                        capture_output=True)
    assert r3.returncode == 0, "--force должен разрешать перезапись"


def test_gsheet_format_emits_bridge(tmp_path):
    out = tmp_path / "g.xlsx"
    r = subprocess.run([sys.executable, str(SCRIPT), "--out", str(out),
                        "--city", "Тест", "--title", "Объект А", "--format", "gsheet"],
                       capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    assert out.exists(), "в gsheet-режиме .xlsx всё равно должен собираться"
    guide = out.with_name(out.stem + " — Google-Таблица (инструкция).md")
    assert guide.exists(), "не создан файл-инструкция (мост) для Google Таблиц"
    txt = guide.read_text(encoding="utf-8")
    assert "Матрица ролей" in txt and "Прораб" in txt, "в инструкции нет матрицы ролей"
    assert ("Защити" in txt or "ЗАЩИТИ" in txt or "Защищённые" in txt), "нет шага про защиту формул"
    assert "A2" in txt, "нет чек-листа проверки переноса A2"
